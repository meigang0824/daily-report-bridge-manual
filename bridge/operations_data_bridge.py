#!/usr/bin/env python3
"""Read-only data service for Dify operations review workflows."""

from __future__ import annotations

import os
import re
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

import pymysql
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook

try:
    import xlrd
except ImportError:
    xlrd = None


DB_CONFIG = {
    "host": os.environ.get("SM_DATA_SQL_HOST", ""),
    "port": int(os.environ.get("SM_DATA_SQL_PORT", "3306")),
    "user": os.environ.get("SM_DATA_SQL_USER", ""),
    "password": os.environ.get("SM_DATA_SQL_PASSWORD", ""),
    "database": os.environ.get("SM_DATA_SQL_DATABASE", ""),
    "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor,
}
REQUIRED_DB_VARS = [
    "SM_DATA_SQL_HOST",
    "SM_DATA_SQL_PORT",
    "SM_DATA_SQL_USER",
    "SM_DATA_SQL_PASSWORD",
    "SM_DATA_SQL_DATABASE",
]

app = FastAPI(
    title="Operations Review Data Bridge",
    description="Read-only structured data API for Dify operations review workflows.",
    version="1.0.0",
)


def require_db_config() -> None:
    missing = [name for name in REQUIRED_DB_VARS if not os.environ.get(name)]
    if missing:
        raise RuntimeError(f"数据库配置不完整: {', '.join(missing)}")


def parse_date(value: str, label: str) -> date:
    try:
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    except (AttributeError, ValueError) as exc:
        raise ValueError(f"{label}格式错误，请填写 YYYY-MM-DD") from exc


def parse_operator_name(filename: str) -> str:
    match = re.match(r"^(.+)-负责店铺列表\.(?:xls|xlsx)$", Path(filename).name, re.IGNORECASE)
    if not match or not match.group(1).strip():
        raise ValueError("店铺列表文件名必须为：{运营姓名}-负责店铺列表.xlsx")
    return match.group(1).strip()


async def read_shop_list(upload: UploadFile) -> tuple[str, List[str]]:
    if not upload.filename or Path(upload.filename).suffix.lower() not in {".xls", ".xlsx"}:
        raise ValueError("负责店铺列表必须为 .xls 或 .xlsx 文件")
    operator_name = parse_operator_name(upload.filename)
    data = await upload.read()
    if len(data) > 10 * 1024 * 1024:
        raise ValueError("负责店铺列表文件超过 10 MB 限制")
    suffix = Path(upload.filename).suffix.lower()
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as target:
        target.write(data)
        temporary_path = Path(target.name)
    try:
        rows = []
        if suffix == ".xls":
            if xlrd is None:
                raise ValueError("读取 .xls 文件需要安装 xlrd")
            book = xlrd.open_workbook(str(temporary_path))
            sheet = book.sheet_by_index(0)
            rows = [sheet.row_values(row) for row in range(sheet.nrows)]
        else:
            workbook = load_workbook(str(temporary_path), data_only=True)
            rows = [list(row) for row in workbook.active.iter_rows(values_only=True)]
            workbook.close()
    finally:
        temporary_path.unlink(missing_ok=True)
    skipped = {"店铺", "店铺名", "店铺名称", "shop", "shop_name"}
    shops = sorted(
        {
            str(row[0]).strip()
            for row in rows
            if row and row[0] is not None and str(row[0]).strip().lower() not in skipped
        }
    )
    if not shops:
        raise ValueError("负责店铺列表中没有解析到店铺名称")
    return operator_name, shops


def number(value: Any) -> float:
    return round(float(value or 0), 2)


def rate(numerator: float, denominator: float) -> Optional[float]:
    if not denominator:
        return None
    return round(numerator / denominator * 100, 2)


def change(current: float, previous: float) -> Dict[str, Optional[float]]:
    return {
        "current": round(current, 2),
        "previous": round(previous, 2),
        "delta": round(current - previous, 2),
        "change_rate": rate(current - previous, previous),
    }


def ratio_change(current: Optional[float], previous: Optional[float]) -> Dict[str, Optional[float]]:
    return {
        "current": current,
        "previous": previous,
        "delta_pp": round(current - previous, 2) if current is not None and previous is not None else None,
    }


def placeholders(values: List[str]) -> str:
    return ", ".join(["%s"] * len(values))


def rows_by_shop(
    connection: pymysql.connections.Connection,
    shops: List[str],
    start: date,
    end: date,
) -> Dict[str, Dict[str, Any]]:
    shop_slots = placeholders(shops)
    params = [start, end, *shops]
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                mall_name,
                COUNT(DISTINCT report_date) AS data_days,
                SUM(COALESCE(goods_visitor_count, 0)) AS visitors,
                SUM(COALESCE(order_count, 0)) AS orders,
                SUM(COALESCE(order_amount, 0)) AS sales,
                SUM(COALESCE(refund_amount, 0)) AS refunds,
                SUM(COALESCE(order_amount, 0) - COALESCE(refund_amount, 0)) AS net_sales,
                SUM(COALESCE(promotion_fee, 0)) AS promotion_fee
            FROM pdd_mall_daily_performance
            WHERE report_date BETWEEN %s AND %s
              AND mall_name IN ({shop_slots})
            GROUP BY mall_name
            """,
            params,
        )
        output: Dict[str, Dict[str, Any]] = {}
        for row in cursor.fetchall():
            output[row["mall_name"]] = {
                "data_days": int(row["data_days"] or 0),
                "visitors": number(row["visitors"]),
                "orders": number(row["orders"]),
                "sales": number(row["sales"]),
                "refunds": number(row["refunds"]),
                "net_sales": number(row["net_sales"]),
                "promotion_fee": number(row["promotion_fee"]),
            }
        cursor.execute(
            f"""
            SELECT
                mall_name,
                SUM(
                    CASE
                        WHEN accounting_type IN ('其他服务', '扣款', '技术服务费', '其他', '多多进宝', '分账')
                        THEN ABS(COALESCE(expense_amount, 0))
                        ELSE 0
                    END
                ) AS operating_expense
            FROM pdd_mall_daily_summary
            WHERE DATE(occurrence_time) BETWEEN %s AND %s
              AND mall_name IN ({shop_slots})
            GROUP BY mall_name
            """,
            params,
        )
        for row in cursor.fetchall():
            output.setdefault(row["mall_name"], {})["operating_expense"] = number(row["operating_expense"])
        cursor.execute(
            f"""
            SELECT shop_name, SUM(COALESCE(today_service_fee, 0)) AS service_fee
            FROM pdd_mall_service_fee_stats
            WHERE stat_date BETWEEN %s AND %s
              AND shop_name IN ({shop_slots})
            GROUP BY shop_name
            """,
            params,
        )
        for row in cursor.fetchall():
            output.setdefault(row["shop_name"], {})["service_fee"] = number(row["service_fee"])
    for shop in shops:
        values = output.setdefault(shop, {})
        for field in ("visitors", "orders", "sales", "refunds", "net_sales", "promotion_fee", "operating_expense", "service_fee"):
            values.setdefault(field, 0.0)
        values.setdefault("data_days", 0)
        values["conversion_rate"] = rate(values["orders"], values["visitors"])
        values["average_order_value"] = round(values["sales"] / values["orders"], 2) if values["orders"] else None
        values["refund_rate"] = rate(values["refunds"], values["sales"])
        values["promotion_rate"] = rate(values["promotion_fee"], values["net_sales"])
        values["operating_expense_rate"] = rate(values["operating_expense"], values["net_sales"])
    return output


def totals(by_shop: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    fields = ("visitors", "orders", "sales", "refunds", "net_sales", "promotion_fee", "operating_expense", "service_fee")
    result = {field: round(sum(row[field] for row in by_shop.values()), 2) for field in fields}
    result["conversion_rate"] = rate(result["orders"], result["visitors"])
    result["average_order_value"] = round(result["sales"] / result["orders"], 2) if result["orders"] else None
    result["refund_rate"] = rate(result["refunds"], result["sales"])
    result["promotion_rate"] = rate(result["promotion_fee"], result["net_sales"])
    result["operating_expense_rate"] = rate(result["operating_expense"], result["net_sales"])
    return result


def compare_metrics(current: Dict[str, Any], previous: Dict[str, Any]) -> Dict[str, Any]:
    result = {}
    for field in ("visitors", "orders", "sales", "refunds", "net_sales", "promotion_fee", "operating_expense", "service_fee"):
        result[field] = change(current[field], previous[field])
    for field in ("conversion_rate", "average_order_value", "refund_rate", "promotion_rate", "operating_expense_rate"):
        if field == "average_order_value":
            curr = current[field] or 0
            prev = previous[field] or 0
            result[field] = change(curr, prev)
        else:
            result[field] = ratio_change(current[field], previous[field])
    return result


def shop_comparisons(
    shops: List[str],
    current: Dict[str, Dict[str, Any]],
    previous: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    rows = []
    for shop in shops:
        rows.append(
            {
                "shop_name": shop,
                "current": current[shop],
                "previous": previous[shop],
                "comparison": compare_metrics(current[shop], previous[shop]),
            }
        )
    return sorted(rows, key=lambda item: item["current"]["net_sales"], reverse=True)


def daily_trend(
    connection: pymysql.connections.Connection,
    shops: List[str],
    start: date,
    end: date,
) -> List[Dict[str, Any]]:
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                report_date,
                SUM(COALESCE(goods_visitor_count, 0)) AS visitors,
                SUM(COALESCE(order_count, 0)) AS orders,
                SUM(COALESCE(order_amount, 0) - COALESCE(refund_amount, 0)) AS net_sales,
                SUM(COALESCE(promotion_fee, 0)) AS promotion_fee
            FROM pdd_mall_daily_performance
            WHERE report_date BETWEEN %s AND %s
              AND mall_name IN ({placeholders(shops)})
            GROUP BY report_date
            ORDER BY report_date
            """,
            [start, end, *shops],
        )
        return [
            {
                "date": row["report_date"].isoformat(),
                "visitors": number(row["visitors"]),
                "orders": number(row["orders"]),
                "net_sales": number(row["net_sales"]),
                "promotion_fee": number(row["promotion_fee"]),
                "conversion_rate": rate(number(row["orders"]), number(row["visitors"])),
                "promotion_rate": rate(number(row["promotion_fee"]), number(row["net_sales"])),
            }
            for row in cursor.fetchall()
        ]


def limited(rows: List[Dict[str, Any]], limit: int) -> List[Dict[str, Any]]:
    return rows[:limit]


def product_rows(
    connection: pymysql.connections.Connection,
    shops: List[str],
    current_start: date,
    current_end: date,
    previous_start: date,
    previous_end: date,
    limit: int = 15,
) -> List[Dict[str, Any]]:
    result: Dict[tuple[str, str], Dict[str, Any]] = {}
    for label, start, end in (("current", current_start, current_end), ("previous", previous_start, previous_end)):
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT mall_name, product_id, MAX(product_title) product_title,
                    SUM(COALESCE(goods_uv, 0)) visitors,
                    SUM(COALESCE(goods_pv, 0)) views,
                    SUM(COALESCE(impr_usr_cnt, 0)) impressions,
                    SUM(COALESCE(pay_ordr_cnt, 0)) orders,
                    SUM(COALESCE(pay_ordr_amt, 0)) sales,
                    SUM(COALESCE(goods_fav_cnt, 0)) favorites,
                    SUM(COALESCE(cnslt_usr_qty, 0)) consultations
                FROM pdd_product_daily_info_stats
                WHERE stat_date BETWEEN %s AND %s
                  AND mall_name IN ({placeholders(shops)})
                GROUP BY mall_name, product_id
                """,
                [start, end, *shops],
            )
            for row in cursor.fetchall():
                key = (row["mall_name"], row["product_id"])
                item = result.setdefault(
                    key,
                    {"mall_name": row["mall_name"], "product_id": row["product_id"], "product_title": row["product_title"]},
                )
                values = {
                    "visitors": number(row["visitors"]),
                    "views": number(row["views"]),
                    "impressions": number(row["impressions"]),
                    "orders": number(row["orders"]),
                    "sales": number(row["sales"]),
                    "favorites": number(row["favorites"]),
                    "consultations": number(row["consultations"]),
                }
                values["conversion_rate"] = rate(values["orders"], values["visitors"])
                item[label] = values
    blank = {
        "visitors": 0.0,
        "views": 0.0,
        "impressions": 0.0,
        "orders": 0.0,
        "sales": 0.0,
        "favorites": 0.0,
        "consultations": 0.0,
        "conversion_rate": None,
    }
    output = []
    for item in result.values():
        item.setdefault("current", blank.copy())
        item.setdefault("previous", blank.copy())
        item["comparison"] = {
            "sales": change(item["current"]["sales"], item["previous"]["sales"]),
            "orders": change(item["current"]["orders"], item["previous"]["orders"]),
            "visitors": change(item["current"]["visitors"], item["previous"]["visitors"]),
            "conversion_rate": ratio_change(item["current"]["conversion_rate"], item["previous"]["conversion_rate"]),
        }
        output.append(item)
    return sorted(output, key=lambda item: item["current"]["sales"], reverse=True)[:limit]


def product_insights(rows: List[Dict[str, Any]], limit: int = 10) -> Dict[str, List[Dict[str, Any]]]:
    active = [row for row in rows if row["current"]["sales"] > 0 or row["previous"]["sales"] > 0]
    return {
        "top_sales": limited(sorted(active, key=lambda item: item["current"]["sales"], reverse=True), limit),
        "top_orders": limited(sorted(active, key=lambda item: item["current"]["orders"], reverse=True), limit),
        "top_visitors": limited(sorted(active, key=lambda item: item["current"]["visitors"], reverse=True), limit),
        "sales_decline": limited(sorted(active, key=lambda item: item["comparison"]["sales"]["delta"]), limit),
        "traffic_decline": limited(sorted(active, key=lambda item: item["comparison"]["visitors"]["delta"]), limit),
        "high_traffic_low_conversion": limited(
            sorted(
                [
                    row
                    for row in active
                    if row["current"]["visitors"] >= 100
                    and (row["current"]["conversion_rate"] is None or row["current"]["conversion_rate"] < 5)
                ],
                key=lambda item: item["current"]["visitors"],
                reverse=True,
            ),
            limit,
        ),
        "new_sales_products": limited(
            sorted(
                [row for row in active if row["current"]["sales"] > 0 and row["previous"]["sales"] == 0],
                key=lambda item: item["current"]["sales"],
                reverse=True,
            ),
            limit,
        ),
        "lost_sales_products": limited(
            sorted(
                [row for row in active if row["current"]["sales"] == 0 and row["previous"]["sales"] > 0],
                key=lambda item: item["previous"]["sales"],
                reverse=True,
            ),
            limit,
        ),
    }


def advertising_rows(
    connection: pymysql.connections.Connection,
    shops: List[str],
    current_start: date,
    current_end: date,
    previous_start: date,
    previous_end: date,
    limit: int = 15,
) -> List[Dict[str, Any]]:
    result: Dict[tuple[str, str], Dict[str, Any]] = {}
    for label, start, end in (("current", current_start, current_end), ("previous", previous_start, previous_end)):
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT mall_name, product_id, MAX(product_title) product_title,
                    SUM(COALESCE(total_cost, 0)) cost,
                    SUM(COALESCE(net_transaction_amount, 0)) net_transaction_amount,
                    SUM(COALESCE(net_transaction_count, 0)) net_transaction_count,
                    SUM(COALESCE(impressions, 0)) impressions,
                    SUM(COALESCE(clicks, 0)) clicks
                FROM pdd_product_daily_advertising_stats
                WHERE report_date BETWEEN %s AND %s
                  AND mall_name IN ({placeholders(shops)})
                GROUP BY mall_name, product_id
                """,
                [start, end, *shops],
            )
            for row in cursor.fetchall():
                key = (row["mall_name"], row["product_id"])
                item = result.setdefault(
                    key,
                    {"mall_name": row["mall_name"], "product_id": row["product_id"], "product_title": row["product_title"]},
                )
                values = {
                    "cost": number(row["cost"]),
                    "net_transaction_amount": number(row["net_transaction_amount"]),
                    "net_transaction_count": number(row["net_transaction_count"]),
                    "impressions": number(row["impressions"]),
                    "clicks": number(row["clicks"]),
                }
                values["roi"] = round(values["net_transaction_amount"] / values["cost"], 2) if values["cost"] else None
                values["ctr"] = rate(values["clicks"], values["impressions"])
                item[label] = values
    blank = {
        "cost": 0.0,
        "net_transaction_amount": 0.0,
        "net_transaction_count": 0.0,
        "impressions": 0.0,
        "clicks": 0.0,
        "roi": None,
        "ctr": None,
    }
    output = []
    for item in result.values():
        item.setdefault("current", blank.copy())
        item.setdefault("previous", blank.copy())
        item["comparison"] = {
            "cost": change(item["current"]["cost"], item["previous"]["cost"]),
            "net_transaction_amount": change(item["current"]["net_transaction_amount"], item["previous"]["net_transaction_amount"]),
            "roi": {
                "current": item["current"]["roi"],
                "previous": item["previous"]["roi"],
                "delta": round(item["current"]["roi"] - item["previous"]["roi"], 2)
                if item["current"]["roi"] is not None and item["previous"]["roi"] is not None
                else None,
            },
        }
        output.append(item)
    return sorted(output, key=lambda item: item["current"]["cost"], reverse=True)[:limit]


def advertising_insights(rows: List[Dict[str, Any]], limit: int = 10) -> Dict[str, List[Dict[str, Any]]]:
    active = [row for row in rows if row["current"]["cost"] > 0 or row["previous"]["cost"] > 0]
    return {
        "top_cost": limited(sorted(active, key=lambda item: item["current"]["cost"], reverse=True), limit),
        "top_transaction_amount": limited(
            sorted(active, key=lambda item: item["current"]["net_transaction_amount"], reverse=True),
            limit,
        ),
        "low_roi": limited(
            sorted(
                [
                    row
                    for row in active
                    if row["current"]["cost"] >= 100
                    and (row["current"]["roi"] is None or row["current"]["roi"] < 1)
                ],
                key=lambda item: item["current"]["cost"],
                reverse=True,
            ),
            limit,
        ),
        "roi_decline": limited(
            sorted(
                [row for row in active if row["comparison"]["roi"]["delta"] is not None],
                key=lambda item: item["comparison"]["roi"]["delta"],
            ),
            limit,
        ),
        "cost_increase": limited(
            sorted(active, key=lambda item: item["comparison"]["cost"]["delta"], reverse=True),
            limit,
        ),
        "high_click_low_conversion": limited(
            sorted(
                [
                    row
                    for row in active
                    if row["current"]["clicks"] >= 50 and row["current"]["net_transaction_count"] <= 3
                ],
                key=lambda item: item["current"]["clicks"],
                reverse=True,
            ),
            limit,
        ),
    }


def expense_breakdown(
    connection: pymysql.connections.Connection,
    shops: List[str],
    current_start: date,
    current_end: date,
    previous_start: date,
    previous_end: date,
) -> Dict[str, Any]:
    values: Dict[str, Dict[str, Dict[str, float]]] = {}
    for label, start, end in (("current", current_start, current_end), ("previous", previous_start, previous_end)):
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT accounting_type, SUM(ABS(COALESCE(expense_amount, 0))) amount, COUNT(*) count
                FROM pdd_mall_daily_summary
                WHERE DATE(occurrence_time) BETWEEN %s AND %s
                  AND mall_name IN ({placeholders(shops)})
                GROUP BY accounting_type
                """,
                [start, end, *shops],
            )
            for row in cursor.fetchall():
                item = values.setdefault(row["accounting_type"] or "未分类", {})
                item[label] = {"amount": number(row["amount"]), "count": int(row["count"] or 0)}
    output = []
    for accounting_type, item in values.items():
        current = item.get("current", {"amount": 0.0, "count": 0})
        previous = item.get("previous", {"amount": 0.0, "count": 0})
        output.append(
            {
                "accounting_type": accounting_type,
                "current": current,
                "previous": previous,
                "comparison": {
                    "amount": change(current["amount"], previous["amount"]),
                    "count": change(float(current["count"]), float(previous["count"])),
                },
            }
        )
    operating_types = {"其他服务", "扣款", "技术服务费", "其他", "多多进宝", "分账"}
    return {
        "by_accounting_type": sorted(output, key=lambda row: row["current"]["amount"], reverse=True),
        "operating_expense_types": sorted(
            [row for row in output if row["accounting_type"] in operating_types],
            key=lambda row: row["current"]["amount"],
            reverse=True,
        ),
        "refund_flow": next((row for row in output if row["accounting_type"] == "退款"), None),
    }


def traffic_funnel(
    connection: pymysql.connections.Connection,
    shops: List[str],
    current_start: date,
    current_end: date,
    previous_start: date,
    previous_end: date,
) -> Dict[str, Any]:
    def query(start: date, end: date) -> Dict[str, Any]:
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT
                    SUM(COALESCE(impr_usr_cnt, 0)) impressions,
                    SUM(COALESCE(goods_uv, 0)) product_visitors,
                    SUM(COALESCE(goods_pv, 0)) product_views,
                    SUM(COALESCE(pay_ordr_cnt, 0)) product_orders,
                    SUM(COALESCE(pay_ordr_amt, 0)) product_sales,
                    SUM(COALESCE(goods_fav_cnt, 0)) favorites,
                    SUM(COALESCE(cnslt_usr_qty, 0)) consultations
                FROM pdd_product_daily_info_stats
                WHERE stat_date BETWEEN %s AND %s
                  AND mall_name IN ({placeholders(shops)})
                """,
                [start, end, *shops],
            )
            row = cursor.fetchone()
        result = {key: number(row[key]) for key in row}
        result["product_conversion_rate"] = rate(result["product_orders"], result["product_visitors"])
        result["average_product_order_value"] = (
            round(result["product_sales"] / result["product_orders"], 2) if result["product_orders"] else None
        )
        result["favorite_rate"] = rate(result["favorites"], result["product_visitors"])
        result["consultation_rate"] = rate(result["consultations"], result["product_visitors"])
        return result

    current = query(current_start, current_end)
    previous = query(previous_start, previous_end)
    comparison = {}
    for field in ("impressions", "product_visitors", "product_views", "product_orders", "product_sales", "favorites", "consultations"):
        comparison[field] = change(current[field], previous[field])
    for field in ("product_conversion_rate", "average_product_order_value", "favorite_rate", "consultation_rate"):
        if field == "average_product_order_value":
            comparison[field] = change(current[field] or 0, previous[field] or 0)
        else:
            comparison[field] = ratio_change(current[field], previous[field])
    return {"current": current, "previous": previous, "comparison": comparison}


def contribution_rankings(shop_rows: List[Dict[str, Any]], limit: int = 10) -> Dict[str, List[Dict[str, Any]]]:
    def project(row: Dict[str, Any], field: str) -> Dict[str, Any]:
        return {
            "shop_name": row["shop_name"],
            "current": row["current"][field],
            "previous": row["previous"][field],
            "delta": row["comparison"][field].get("delta"),
            "change_rate": row["comparison"][field].get("change_rate"),
        }

    rankings = {}
    for field in ("net_sales", "orders", "visitors", "promotion_fee", "refunds", "operating_expense"):
        projected = [project(row, field) for row in shop_rows]
        rankings[f"{field}_growth_top"] = limited(sorted(projected, key=lambda item: item["delta"], reverse=True), limit)
        rankings[f"{field}_decline_top"] = limited(sorted(projected, key=lambda item: item["delta"]), limit)
    return rankings


def refund_insights(shop_rows: List[Dict[str, Any]], expense: Dict[str, Any], limit: int = 10) -> Dict[str, Any]:
    refund_rows = [
        {
            "shop_name": row["shop_name"],
            "refunds": row["comparison"]["refunds"],
            "refund_rate": row["comparison"]["refund_rate"],
            "current_sales": row["current"]["sales"],
        }
        for row in shop_rows
    ]
    return {
        "refund_flow_by_accounting_type": expense.get("refund_flow"),
        "refund_amount_top": limited(
            sorted(refund_rows, key=lambda item: item["refunds"]["current"], reverse=True),
            limit,
        ),
        "refund_increase_top": limited(
            sorted(refund_rows, key=lambda item: item["refunds"]["delta"], reverse=True),
            limit,
        ),
        "high_refund_rate": limited(
            sorted(
                [row for row in refund_rows if row["refund_rate"]["current"] is not None and row["refund_rate"]["current"] >= 15],
                key=lambda item: item["refund_rate"]["current"],
                reverse=True,
            ),
            limit,
        ),
    }


def score_insights(score_rows_data: List[Dict[str, Any]], limit: int = 10) -> Dict[str, Any]:
    return {
        "score_decline": limited(
            sorted(
                [row for row in score_rows_data if row["score_delta"] is not None],
                key=lambda item: item["score_delta"],
            ),
            limit,
        ),
        "low_current_score": limited(
            sorted(
                [row for row in score_rows_data if row.get("current")],
                key=lambda item: item["current"]["total_score"],
            ),
            limit,
        ),
    }


def product_score_insights(
    connection: pymysql.connections.Connection,
    shops: List[str],
    current_end: date,
    previous_end: date,
    limit: int = 10,
) -> Dict[str, List[Dict[str, Any]]]:
    result: Dict[tuple[str, str], Dict[str, Any]] = {}
    for label, period_end in (("current", current_end), ("previous", previous_end)):
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT mall_name, product_id, MAX(product_title) product_title,
                       SUM(COALESCE(review_total_cnt, 0)) review_total_cnt,
                       AVG(COALESCE(ranking_score, 0)) ranking_score
                FROM pdd_product_daily_dsr_stats
                WHERE stat_date = %s
                  AND mall_name IN ({placeholders(shops)})
                GROUP BY mall_name, product_id
                """,
                [period_end, *shops],
            )
            for row in cursor.fetchall():
                key = (row["mall_name"], row["product_id"])
                item = result.setdefault(
                    key,
                    {"mall_name": row["mall_name"], "product_id": row["product_id"], "product_title": row["product_title"]},
                )
                item[label] = {
                    "review_total_cnt": number(row["review_total_cnt"]),
                    "ranking_score": number(row["ranking_score"]),
                }
    rows = []
    for item in result.values():
        current = item.get("current")
        previous = item.get("previous")
        item["score_delta"] = (
            round(current["ranking_score"] - previous["ranking_score"], 2) if current and previous else None
        )
        if current or previous:
            rows.append(item)
    return {
        "product_score_decline": limited(
            sorted([row for row in rows if row["score_delta"] is not None], key=lambda item: item["score_delta"]),
            limit,
        ),
        "low_score_high_review_products": limited(
            sorted(
                [
                    row
                    for row in rows
                    if row.get("current")
                    and row["current"]["review_total_cnt"] >= 10
                    and row["current"]["ranking_score"] <= 4.6
                ],
                key=lambda item: (item["current"]["ranking_score"], -item["current"]["review_total_cnt"]),
            ),
            limit,
        ),
    }


def build_alerts(
    summary: Dict[str, Any],
    shop_rows: List[Dict[str, Any]],
    product_insight_data: Dict[str, List[Dict[str, Any]]],
    ad_insight_data: Dict[str, List[Dict[str, Any]]],
    warnings: List[str],
) -> List[Dict[str, Any]]:
    alerts = [{"level": "warning", "type": "data_coverage", "message": warning} for warning in warnings]
    if summary["comparison"]["net_sales"]["change_rate"] is not None and summary["comparison"]["net_sales"]["change_rate"] <= -20:
        alerts.append({"level": "high", "type": "summary_net_sales_drop", "message": "整体净销售额环比下降超过 20%", "data": summary["comparison"]["net_sales"]})
    if summary["comparison"]["orders"]["change_rate"] is not None and summary["comparison"]["orders"]["change_rate"] <= -20:
        alerts.append({"level": "high", "type": "summary_orders_drop", "message": "整体订单量环比下降超过 20%", "data": summary["comparison"]["orders"]})
    if summary["current"]["refund_rate"] is not None and summary["current"]["refund_rate"] >= 15:
        alerts.append({"level": "high", "type": "summary_high_refund_rate", "message": "整体退款率达到 15% 及以上", "data": summary["current"]["refund_rate"]})
    if summary["current"]["promotion_rate"] is not None and summary["current"]["promotion_rate"] >= 35:
        alerts.append({"level": "medium", "type": "summary_high_promotion_rate", "message": "整体推广费率达到 35% 及以上", "data": summary["current"]["promotion_rate"]})

    for row in shop_rows:
        shop = row["shop_name"]
        if row["comparison"]["net_sales"]["change_rate"] is not None and row["comparison"]["net_sales"]["change_rate"] <= -30:
            alerts.append({"level": "high", "type": "shop_net_sales_drop", "message": f"{shop} 净销售额下降超过 30%", "data": row["comparison"]["net_sales"]})
        if row["current"]["refund_rate"] is not None and row["current"]["refund_rate"] >= 20:
            alerts.append({"level": "high", "type": "shop_high_refund_rate", "message": f"{shop} 退款率达到 20% 及以上", "data": row["current"]["refund_rate"]})
        if row["current"]["promotion_rate"] is not None and row["current"]["promotion_rate"] >= 45:
            alerts.append({"level": "medium", "type": "shop_high_promotion_rate", "message": f"{shop} 推广费率达到 45% 及以上", "data": row["current"]["promotion_rate"]})

    for item in ad_insight_data["low_roi"][:5]:
        alerts.append({"level": "medium", "type": "ad_low_roi", "message": "高花费低 ROI 投放商品", "data": item})
    for item in product_insight_data["high_traffic_low_conversion"][:5]:
        alerts.append({"level": "medium", "type": "product_high_traffic_low_conversion", "message": "高流量低转化商品", "data": item})
    return alerts


def score_rows(
    connection: pymysql.connections.Connection,
    shops: List[str],
    current_end: date,
    previous_end: date,
) -> List[Dict[str, Any]]:
    output: Dict[str, Dict[str, Any]] = {shop: {"shop_name": shop} for shop in shops}
    for label, period_end in (("current", current_end), ("previous", previous_end)):
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT scores.mall_name, scores.stat_date, scores.total_score, scores.attitude_score,
                       scores.product_score, scores.delivery_score, scores.logistics_score
                FROM pdd_mall_consumer_scores_daily scores
                INNER JOIN (
                    SELECT mall_name, MAX(stat_date) stat_date
                    FROM pdd_mall_consumer_scores_daily
                    WHERE stat_date <= %s AND mall_name IN ({placeholders(shops)})
                    GROUP BY mall_name
                ) latest
                  ON latest.mall_name = scores.mall_name AND latest.stat_date = scores.stat_date
                """,
                [period_end, *shops],
            )
            for row in cursor.fetchall():
                output[row["mall_name"]][label] = {
                    "stat_date": row["stat_date"].isoformat(),
                    "total_score": number(row["total_score"]),
                    "attitude_score": number(row["attitude_score"]),
                    "product_score": number(row["product_score"]),
                    "delivery_score": number(row["delivery_score"]),
                    "logistics_score": number(row["logistics_score"]),
                }
    result = []
    for item in output.values():
        current = item.get("current")
        previous = item.get("previous")
        item["score_delta"] = (
            round(current["total_score"] - previous["total_score"], 2) if current and previous else None
        )
        if current or previous:
            result.append(item)
    return result


def data_coverage(
    connection: pymysql.connections.Connection,
    shops: List[str],
    current_start: date,
    current_end: date,
    previous_start: date,
    previous_end: date,
) -> Dict[str, Any]:
    with connection.cursor() as cursor:
        cursor.execute("SELECT MIN(report_date) min_date, MAX(report_date) max_date FROM pdd_mall_daily_performance")
        bounds = cursor.fetchone()
        cursor.execute(
            f"""
            SELECT DISTINCT mall_name
            FROM pdd_mall_daily_performance
            WHERE mall_name IN ({placeholders(shops)})
              AND report_date BETWEEN %s AND %s
            """,
            [*shops, min(current_start, previous_start), max(current_end, previous_end)],
        )
        matched = sorted(row["mall_name"] for row in cursor.fetchall())
    missing = sorted(set(shops) - set(matched))
    return {
        "performance_available_range": {
            "min_date": bounds["min_date"].isoformat() if bounds["min_date"] else None,
            "max_date": bounds["max_date"].isoformat() if bounds["max_date"] else None,
        },
        "matched_shops": matched,
        "unmatched_shops": missing,
        "requested_shop_count": len(shops),
    }


def build_comparison_payload(
    analysis_type: str,
    operator_name: str,
    shops: List[str],
    current_start: date,
    current_end: date,
    previous_start: date,
    previous_end: date,
) -> Dict[str, Any]:
    require_db_config()
    connection = pymysql.connect(**DB_CONFIG)
    try:
        current = rows_by_shop(connection, shops, current_start, current_end)
        previous = rows_by_shop(connection, shops, previous_start, previous_end)
        current_totals = totals(current)
        previous_totals = totals(previous)
        coverage = data_coverage(connection, shops, current_start, current_end, previous_start, previous_end)
        warnings = []
        if coverage["unmatched_shops"]:
            warnings.append("以下名单店铺在请求周期内没有匹配到经营数据：" + "、".join(coverage["unmatched_shops"]))
        expected_days = (current_end - current_start).days + 1
        current_missing = [shop for shop in shops if current[shop]["data_days"] < expected_days]
        previous_missing = [shop for shop in shops if previous[shop]["data_days"] < expected_days]
        if current_missing:
            warnings.append("本期数据天数不完整的店铺：" + "、".join(current_missing))
        if previous_missing:
            warnings.append("对比期数据天数不完整的店铺：" + "、".join(previous_missing))
        summary = {
            "current": current_totals,
            "previous": previous_totals,
            "comparison": compare_metrics(current_totals, previous_totals),
        }
        shops_payload = shop_comparisons(shops, current, previous)
        product_rows_payload = product_rows(
            connection, shops, current_start, current_end, previous_start, previous_end, limit=100000
        )
        advertising_rows_payload = advertising_rows(
            connection, shops, current_start, current_end, previous_start, previous_end, limit=100000
        )
        product_insight_payload = product_insights(product_rows_payload)
        advertising_insight_payload = advertising_insights(advertising_rows_payload)
        expense_payload = expense_breakdown(connection, shops, current_start, current_end, previous_start, previous_end)
        score_payload = score_rows(connection, shops, current_end, previous_end)
        return {
            "success": True,
            "analysis_type": analysis_type,
            "operator_name": operator_name,
            "managed_shops": shops,
            "period": {
                "current_start": current_start.isoformat(),
                "current_end": current_end.isoformat(),
                "previous_start": previous_start.isoformat(),
                "previous_end": previous_end.isoformat(),
            },
            "summary": summary,
            "shop_comparisons": shops_payload,
            "contribution_rankings": contribution_rankings(shops_payload),
            "daily_trend": {
                "current": daily_trend(connection, shops, current_start, current_end),
                "previous": daily_trend(connection, shops, previous_start, previous_end),
            },
            "expense_breakdown": expense_payload,
            "traffic_funnel": traffic_funnel(connection, shops, current_start, current_end, previous_start, previous_end),
            "product_metrics_top_by_current_sales": product_insight_payload["top_sales"],
            "advertising_metrics_top_by_current_cost": advertising_insight_payload["top_cost"],
            "product_insights": product_insight_payload,
            "advertising_insights": advertising_insight_payload,
            "refund_insights": refund_insights(shops_payload, expense_payload),
            "score_metrics": score_payload,
            "score_insights": {
                **score_insights(score_payload),
                **product_score_insights(connection, shops, current_end, previous_end),
            },
            "data_coverage": coverage,
            "warnings": warnings,
            "alerts": build_alerts(summary, shops_payload, product_insight_payload, advertising_insight_payload, warnings),
            "metric_definitions": {
                "net_sales": "成交金额 - 退款金额",
                "conversion_rate": "订单量 / 访客数 * 100%",
                "refund_rate": "退款金额 / 成交金额 * 100%",
                "promotion_rate": "推广费 / 净销售额 * 100%",
                "operating_expense": "其他服务、扣款、技术服务费、其他、多多进宝、分账支出的绝对值合计",
                "operating_expense_rate": "经营支出 / 净销售额 * 100%",
                "change_rate": "(本期 - 对比期) / 对比期 * 100%",
                "traffic_funnel": "来自商品明细表的曝光、商品访客、商品浏览、成交订单、成交金额、收藏和咨询汇总",
                "alerts": "脚本按固定阈值标记的数据异常，只作为 AI 重点解读线索",
            },
        }
    finally:
        connection.close()


async def execute_request(
    analysis_type: str,
    target_date: str,
    shop_list_file: UploadFile,
) -> Dict[str, Any]:
    operator_name, shops = await read_shop_list(shop_list_file)
    end = parse_date(target_date, "复盘日期")
    if analysis_type == "daily_comparison":
        return build_comparison_payload(analysis_type, operator_name, shops, end, end, end - timedelta(days=1), end - timedelta(days=1))
    current_start = end - timedelta(days=6)
    previous_end = current_start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=6)
    return build_comparison_payload(analysis_type, operator_name, shops, current_start, end, previous_start, previous_end)


@app.get("/health")
async def health() -> Dict[str, Any]:
    return {"status": "healthy", "read_only": True, "database_configured": not any(not os.environ.get(name) for name in REQUIRED_DB_VARS)}


@app.post("/v1/analysis/daily-comparison")
async def daily_comparison(
    data_date: str = Form(...),
    shop_list_file: UploadFile = File(...),
):
    try:
        return await execute_request("daily_comparison", data_date, shop_list_file)
    except ValueError as exc:
        return JSONResponse({"success": False, "message": str(exc)}, status_code=200)
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc


@app.post("/v1/analysis/rolling-7d-comparison")
async def rolling_7d_comparison(
    end_date: str = Form(...),
    shop_list_file: UploadFile = File(...),
):
    try:
        return await execute_request("rolling_7d_comparison", end_date, shop_list_file)
    except ValueError as exc:
        return JSONResponse({"success": False, "message": str(exc)}, status_code=200)
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("OPERATIONS_BRIDGE_PORT", "8768")))
