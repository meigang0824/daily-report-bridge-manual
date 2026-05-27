# -*- coding: utf-8 -*-
from fastapi import FastAPI, HTTPException, BackgroundTasks, UploadFile, File, Form, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from typing import Optional, Dict, List, Any
from datetime import datetime, timedelta
import sys
import os
import json
import tempfile
import requests
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl 未安装，请运行: pip install openpyxl")
    sys.exit(1)

try:
    import pymysql
except ImportError:
    print("[ERROR] pymysql 未安装，请运行: pip install pymysql")
    sys.exit(1)

try:
    import xlrd
except ImportError:
    xlrd = None
    print("[WARN] xlrd 未安装，.xls 格式补单文件将无法读取。请运行: pip install xlrd")


app = FastAPI(
    title="每日店铺数据报表 API",
    description="生成每日店铺数据报表的 FastAPI 服务",
    version="2.0.0"
)


def _get_env(name, default=None):
    if sys.platform.startswith('win'):
        try:
            import winreg
            with winreg.OpenKey(winreg.HKEY_CURRENT_USER, 'Environment') as key:
                value, _ = winreg.QueryValueEx(key, name)
                return value
        except (OSError, ImportError):
            pass
    return os.environ.get(name, default)


DB_CONFIG = {
    'host': _get_env('SM_DATA_SQL_HOST', ''),
    'port': int(_get_env('SM_DATA_SQL_PORT', '3306')),
    'user': _get_env('SM_DATA_SQL_USER', ''),
    'password': _get_env('SM_DATA_SQL_PASSWORD', ''),
    'db': _get_env('SM_DATA_SQL_DATABASE', ''),
    'charset': 'utf8mb4'
}

WEKNORA_CONFIG = {
    'api_key': _get_env('WEKNORA_API_KEY', ''),
    'api_url': _get_env('WEKNORA_BASE_URL', ''),
    'knowledge_base_id': '',
    'knowledge_base_name': '张洋的运营知识库',
    'template_file_name': '张洋的店铺数据日报.xlsx',
    'output_file_name': '张洋的店铺数据日报.xlsx',
}

FEISHU_CONFIG = {
    'app_id': '',
    'app_secret': '',
    'receive_id': '',
    'receive_id_type': 'open_id'
}

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = str(BASE_DIR / 'output')
LOG_DIR = str(BASE_DIR / 'logs')


class ReportRequest(BaseModel):
    role: str = Field(default="operator", description="角色: boss/supervisor/operator")
    username: Optional[str] = Field(default="运营主管", description="用户名（运营人员必填）")
    test_date: Optional[str] = Field(default=None, description="测试日期，None表示使用昨天")
    upload_to_knowledge_base: bool = Field(default=True, description="是否上传到知识库")
    send_to_feishu: bool = Field(default=True, description="是否发送到飞书")
    supplement_days_range: int = Field(default=1, description="补单文件查找天数范围（前后各N天）")


class ShopData(BaseModel):
    date: str
    visitor_count: float
    order_count: float
    sales_amount: float
    net_sales: float
    promotion_fee: float
    pay_ratio: float
    net_pay_ratio: float
    expense_amount: float
    daily_expense_rate: float
    service_fee: float
    score_rank: str
    supplement_count: Optional[int] = 0
    supplement_amount: Optional[float] = 0.0
    is_supplement_removed: Optional[bool] = False
    supplement_total_amount: Optional[float] = 0.0
    supplement_order_count: Optional[int] = 0


class ReportResponse(BaseModel):
    success: bool
    message: str
    data_date: str
    shop_count: int
    output_file: Optional[str] = None
    weknora_file_id: Optional[str] = None
    feishu_message_id: Optional[str] = None
    shop_data: Optional[Dict[str, ShopData]] = None


class SupplementUploadResponse(BaseModel):
    success: bool
    message: str
    file_name: str
    weknora_file_id: Optional[str] = None
    report_generated: bool = False
    report_result: Optional[ReportResponse] = None


class BatchSupplementUploadResponse(BaseModel):
    success: bool
    message: str
    total_files: int
    success_count: int
    failed_count: int
    results: List[SupplementUploadResponse]


class ShopListUploadResponse(BaseModel):
    success: bool
    message: str
    shop_count: int
    shops: List[str]
    report_generated: bool = False
    report_result: Optional[ReportResponse] = None


class DailyStoreReportAPI:
    def __init__(self, role: str = "operator", username: Optional[str] = None, test_date: Optional[str] = None, supplement_days_range: int = 1, custom_shops: Optional[List[str]] = None, custom_supplement_data: Optional[Dict[str, Dict]] = None):
        self.log_messages = []
        self.role = role
        self.username = username
        
        # 根据 username 动态生成文件名
        if self.username:
            self.template_file_name = f"{self.username}的店铺数据日报.xlsx"
            self.output_file_name = f"{self.username}的店铺数据日报.xlsx"
        else:
            self.template_file_name = WEKNORA_CONFIG['template_file_name']
            self.output_file_name = WEKNORA_CONFIG['output_file_name']
        
        self.test_date_override = test_date
        self.supplement_days_range = supplement_days_range
        self.custom_supplement_data = custom_supplement_data or {}
        self.is_custom_shops = custom_shops is not None
        self.date_yesterday, self.date_str = self.get_target_date()
        
        if custom_shops:
            self.managed_shops = custom_shops
            self.log(f"[OK] 使用自定义店铺列表，共 {len(custom_shops)} 个店铺")
            self.user_id = None
        else:
            self.user_id = self.get_user_id() if role == 'operator' else None
            self.managed_shops = self.get_managed_shops()

    def log(self, message: str):
        self.log_messages.append(message)
        print(message)

    def get_target_date(self):
        if self.test_date_override:
            self.log(f"\n[测试模式] 使用传入日期: {self.test_date_override}")
            return self.test_date_override, self.test_date_override.replace('-', '')
        yesterday = datetime.now() - timedelta(days=1)
        date = yesterday.strftime('%Y-%m-%d')
        self.log(f"\n[生产模式] 使用昨天日期: {date}")
        return date, yesterday.strftime('%Y%m%d')

    def get_user_id(self):
        if not self.username:
            raise ValueError('[ERROR] 运营人员必须配置 username')
        
        query = f"""
        SELECT id FROM sys_users 
        WHERE username = '{self.username}' AND status = 1
        LIMIT 1
        """
        
        result = self.run_mysql_query(query)
        if not result:
            raise ValueError(f'[ERROR] 未找到用户名为 "{self.username}" 的用户')
        
        user_id = result.strip().split('\t')[0]
        self.log(f"[OK] 找到用户: {self.username} -> {user_id}")
        return user_id

    def get_managed_shops(self):
        if self.role in ['boss', 'supervisor']:
            query = "SELECT shop_name FROM mapping_shops WHERE status = 1 ORDER BY shop_name"
        else:
            query = f"""
            SELECT shop_name FROM mapping_shops 
            WHERE manager_user_id = '{self.user_id}' AND status = 1
            ORDER BY shop_name
            """
        
        result = self.run_mysql_query(query)
        shops = [line.strip() for line in result.split('\n') if line.strip()]
        self.log(f"[OK] 负责的店铺数量: {len(shops)}")
        return shops

    def run_mysql_query(self, query):
        try:
            conn = pymysql.connect(**DB_CONFIG)
            with conn.cursor() as cursor:
                cursor.execute(query)
                rows = cursor.fetchall()
                if not rows:
                    return ''
                lines = []
                for row in rows:
                    lines.append('\t'.join(str(v) if v is not None else '' for v in row))
                return '\n'.join(lines)
        except Exception as e:
            self.log(f"[MySQL Error] {e}")
            return ''
        finally:
            if 'conn' in locals():
                conn.close()

    def get_supplement_removed_status(self, wb):
        supplement_removed_status = {}
        for sheet_name in wb.sheetnames:
            if sheet_name == '店铺单量总表':
                continue
            ws = wb[sheet_name]
            for row in range(3, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                is_removed_val = ws.cell(row=row, column=12).value
                if date_val:
                    key = f"{sheet_name}|{date_val}"
                    supplement_removed_status[key] = bool(is_removed_val) if is_removed_val is not None else False
        return supplement_removed_status

    def query_daily_data(self, supplement_removed_status=None):
        if supplement_removed_status is None:
            supplement_removed_status = {}
        self.log(f"\n[Step 1] 查询 {self.date_yesterday} 的店铺数据...")
        
        if self.role in ['boss', 'supervisor']:
            shop_filter = ''
            summary_shop_filter = ''
        else:
            shop_list = ','.join([f"'{shop}'" for shop in self.managed_shops])
            shop_filter = f"AND p.mall_name IN ({shop_list})"
            summary_shop_filter = f"AND s.shop_name IN ({shop_list})"
        
        query1 = f"""
        SELECT 
            p.mall_name,
            p.report_date,
            COALESCE(p.goods_visitor_count, 0) as visitor_count,
            COALESCE(p.order_count, 0) as order_count,
            COALESCE(p.order_amount, 0) as sales_amount,
            COALESCE(p.order_amount - p.refund_amount, 0) as net_sales,
            COALESCE(p.promotion_fee, 0) as promotion_fee,
            COALESCE(p.pay_ratio, 0) as pay_ratio,
            CASE 
                WHEN (p.order_amount - p.refund_amount) > 0 
                THEN ROUND(p.promotion_fee / (p.order_amount - p.refund_amount) * 100, 2)
                ELSE 0 
            END as net_pay_ratio,
            COALESCE(s.today_service_fee, 0) as service_fee,
            p.mall_score_rank as score_rank
        FROM pdd_mall_daily_performance p
        LEFT JOIN pdd_mall_service_fee_stats s 
            ON p.mall_name = s.shop_name AND p.report_date = s.stat_date
        WHERE p.report_date = '{self.date_yesterday}'
        {shop_filter}
        ORDER BY p.order_amount DESC
        """
        
        result1 = self.run_mysql_query(query1)
        
        daily_data = {}
        for line in result1.split('\n'):
            if line.strip() and '\t' in line:
                fields = line.split('\t')
                if len(fields) >= 11:
                    shop_name = fields[0]
                    daily_data[shop_name] = {
                        'date': fields[1],
                        'visitor_count': float(fields[2]) if fields[2] else 0,
                        'order_count': float(fields[3]) if fields[3] else 0,
                        'sales_amount': float(fields[4]) if fields[4] else 0,
                        'net_sales': float(fields[5]) if fields[5] else 0,
                        'promotion_fee': float(fields[6]) if fields[6] else 0,
                        'pay_ratio': float(fields[7]) if fields[7] else 0,
                        'net_pay_ratio': float(fields[8]) if fields[8] else 0,
                        'expense_amount': 0,
                        'daily_expense_rate': 0,
                        'service_fee': float(fields[9]) if fields[9] else 0,
                        'score_rank': fields[10] if fields[10] else '',
                    }
        
        self.log(f"[OK] 查询到 {len(daily_data)} 个店铺的数据")
        
        self.log(f"[Step 1.5] 查询支出数据（pdd_mall_daily_summary）...")
        
        if self.role in ['boss', 'supervisor']:
            summary_shop_filter = ''
        else:
            shop_list = ','.join([f"'{shop}'" for shop in self.managed_shops])
            summary_shop_filter = f"AND s.mall_name IN ({shop_list})"
        
        query2 = f"""
        SELECT 
            s.mall_name,
            SUM(CASE WHEN s.accounting_type IN ('其他服务', '扣款', '技术服务费', '其他', '多多进宝', '分账')
                THEN ABS(IFNULL(s.expense_amount, 0)) ELSE 0 END) as expense_amount
        FROM pdd_mall_daily_summary s
        WHERE DATE(s.occurrence_time) = '{self.date_yesterday}'
        {summary_shop_filter}
        GROUP BY s.mall_name
        """
        
        result2 = self.run_mysql_query(query2)
        
        for line in result2.split('\n'):
            if line.strip() and '\t' in line:
                fields = line.split('\t')
                if len(fields) >= 2:
                    shop_name = fields[0]
                    expense_amount = float(fields[1]) if fields[1] else 0
                    
                    if shop_name in daily_data:
                        daily_data[shop_name]['expense_amount'] = expense_amount
        
        if self.custom_supplement_data:
            self.log(f"[OK] 使用自定义补单数据，{len(self.custom_supplement_data)} 个店铺有补单")
            supplement_data = self.custom_supplement_data
        else:
            self.log(f"[OK] 未提供自定义补单数据，从知识库查找...")
            supplement_data = self.get_supplement_amounts()

        if supplement_data:
            self.log(f"[OK] 找到补单数据，{len(supplement_data)} 个店铺有补单")
            for shop_name, sdata in supplement_data.items():
                supplement_count = sdata['count']
                supplement_amount = sdata['amount']
                if shop_name in daily_data:
                    key = f"{shop_name}|{self.date_yesterday}"
                    already_removed = supplement_removed_status.get(key, False)
                    
                    if already_removed:
                        self.log(f"[INFO] {shop_name} 在 {self.date_yesterday} 的补单数据已剔除过，跳过重复处理")
                        daily_data[shop_name]['supplement_count'] = supplement_count
                        daily_data[shop_name]['supplement_amount'] = supplement_amount
                        daily_data[shop_name]['is_supplement_removed'] = True
                        daily_data[shop_name]['supplement_total_amount'] = supplement_amount
                        daily_data[shop_name]['supplement_order_count'] = supplement_count
                    else:
                        self.log(f"[INFO] {shop_name} 在 {self.date_yesterday} 首次处理补单数据")
                        daily_data[shop_name]['supplement_count'] = supplement_count
                        daily_data[shop_name]['supplement_amount'] = supplement_amount
                        daily_data[shop_name]['is_supplement_removed'] = False
                        daily_data[shop_name]['supplement_total_amount'] = supplement_amount
                        daily_data[shop_name]['supplement_order_count'] = supplement_count

                        original_order_count = daily_data[shop_name]['order_count']
                        adjusted_order_count = original_order_count - supplement_count
                        daily_data[shop_name]['order_count'] = max(0, adjusted_order_count)

                        original_net_sales = daily_data[shop_name]['net_sales']
                        adjusted_net_sales = original_net_sales - supplement_amount
                        adjusted_net_sales = max(0, adjusted_net_sales)
                        daily_data[shop_name]['net_sales'] = round(adjusted_net_sales, 2)

                        promotion_fee = daily_data[shop_name]['promotion_fee']
                        if adjusted_net_sales > 0:
                            daily_data[shop_name]['net_pay_ratio'] = round(promotion_fee / adjusted_net_sales * 100, 2)
                        else:
                            daily_data[shop_name]['net_pay_ratio'] = 0
        
        self.log(f"[Step 3] 计算支出率（基于补单后净销售额）...")
        for shop_name, data in daily_data.items():
            expense_amount = data.get('expense_amount', 0)
            net_sales = data['net_sales']
            if net_sales > 0 and expense_amount > 0:
                data['daily_expense_rate'] = round(expense_amount / net_sales * 100, 2)
            else:
                data['daily_expense_rate'] = 0
        
        return daily_data

    def _weknora_request(self, method, url, **kwargs):
        headers = kwargs.pop('headers', {})
        headers['X-API-Key'] = WEKNORA_CONFIG['api_key']
        headers['Accept'] = 'application/json'
        kwargs.setdefault('timeout', 30)
        
        _proxy_vars = ['http_proxy', 'https_proxy', 'HTTP_PROXY', 'HTTPS_PROXY', 'ALL_PROXY', 'all_proxy']
        saved = {}
        for var in _proxy_vars:
            if var in os.environ:
                saved[var] = os.environ.pop(var)
        
        try:
            response = requests.request(method, url, headers=headers, **kwargs)
            response.encoding = 'utf-8'
            return response
        except requests.RequestException as e:
            self.log(f"[ERROR] WeKnora 请求失败: {url} -> {e}")
            return None
        finally:
            os.environ.update(saved)

    def get_supplement_amounts(self):
        import re
        self.log(f"\n[Step 2] 查询补单数据...")
        
        list_url = f"{WEKNORA_CONFIG['api_url']}/knowledge-bases/{WEKNORA_CONFIG['knowledge_base_id']}/knowledge"
        response = self._weknora_request('GET', list_url, timeout=10)
        if response is None:
            return None
        
        try:
            result = response.json()
        except Exception as e:
            self.log(f"[WARN] 知识库返回解析失败: {e}")
            return None
        
        if not result.get('success') or not result.get('data'):
            self.log(f"[INFO] 知识库中未找到任何文件")
            return None
        
        self.log(f"[INFO] 知识库中共 {len(result['data'])} 个文件")
        
        date_obj = datetime.strptime(self.date_yesterday, '%Y-%m-%d')
        
        date_patterns = []
        for delta in range(-self.supplement_days_range, self.supplement_days_range + 1):
            d = date_obj + timedelta(days=delta)
            date_patterns.append(d.strftime('%Y-%m-%d'))
            date_patterns.append(f"{d.year}-{d.month}-{d.day}")
        date_patterns = list(set(date_patterns))
        
        self.log(f"[INFO] 查找日期范围: {date_obj - timedelta(days=self.supplement_days_range)} 至 {date_obj + timedelta(days=self.supplement_days_range)}")
        self.log(f"[INFO] 日期匹配模式: {date_patterns}")
        
        file_suffix_re = re.compile(r'单\.(xls|xlsx)$', re.IGNORECASE)
        date_match_re = re.compile(
            r'^(' + '|'.join(re.escape(dp) for dp in date_patterns) + r')\D'
        )
        matched_files = []
        
        for item in result['data']:
            title = item.get('title', '')
            date_matched = bool(date_match_re.match(title))
            suffix_matched = file_suffix_re.search(title)
            if date_matched and suffix_matched:
                matched_files.append((item, title))
                self.log(f"[INFO] 找到补单文件: {title}")
        
        if not matched_files:
            return None
        
        shop_data: dict = defaultdict(lambda: {'count': 0, 'amount': 0.0})
        
        for item, title in matched_files:
            shop_match = re.search(r'^\d{4}-\d{1,2}-\d{1,2}(.+?)\s+\d+单\.\w+$', title)
            if shop_match:
                shop_name = shop_match.group(1).strip()
            else:
                continue
            
            knowledge_id = item['id']
            
            download_url = f"{WEKNORA_CONFIG['api_url']}/knowledge/{knowledge_id}/download"
            dl_response = self._weknora_request('GET', download_url, timeout=30)
            if dl_response is None or dl_response.status_code != 200:
                continue
            
            ext = title.rsplit('.', 1)[-1] if '.' in title else 'xls'
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
            temp_file.write(dl_response.content)
            temp_file.close()
            
            try:
                count, amount = self._parse_supplement_file(temp_file.name, ext)
                shop_data[shop_name]['count'] += count
                shop_data[shop_name]['amount'] += amount
            except Exception as e:
                self.log(f"[WARN] 解析补单文件失败 {title}: {e}")
            finally:
                os.unlink(temp_file.name)
        
        if not shop_data:
            return None
        
        return dict(shop_data)

    def _parse_supplement_file(self, file_path, ext):
        is_xls = ext.lower() == 'xls'
        
        if is_xls and xlrd is None:
            raise RuntimeError("xlrd 未安装，无法读取 .xls 文件")
        
        rows = []
        if is_xls:
            book = xlrd.open_workbook(file_path)
            sheet = book.sheet_by_index(0)
            for r in range(sheet.nrows):
                row = []
                for c in range(sheet.ncols):
                    row.append(sheet.cell_value(r, c))
                rows.append(row)
        else:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()
        
        if not rows:
            return 0, 0.0
        
        header_row_idx = None
        order_col = None
        amount_col = None
        amount_keywords_exact = ['订单金额', '用户实付金额']
        amount_keywords_fuzzy = ['金额']
        
        for idx in range(min(3, len(rows))):
            row = rows[idx]
            str_row = [str(v) if v is not None else '' for v in row]
            if '订单号' in str_row:
                header_row_idx = idx
                order_col = str_row.index('订单号')
                for ci, val in enumerate(str_row):
                    if ci == order_col:
                        continue
                    if any(kw in val for kw in amount_keywords_exact):
                        amount_col = ci
                        break
                if amount_col is None:
                    for ci, val in enumerate(str_row):
                        if ci == order_col:
                            continue
                        if any(kw in val for kw in amount_keywords_fuzzy):
                            amount_col = ci
                            break
                break
        
        if header_row_idx is None or order_col is None:
            return 0, 0.0
        
        if amount_col is None:
            return 0, 0.0
        
        order_count = 0
        total_amount = 0.0
        for r in range(header_row_idx + 1, len(rows)):
            row = rows[r]
            order_id = row[order_col] if order_col < len(row) else None
            if order_id is not None and str(order_id).strip():
                order_count += 1
                amt_val = row[amount_col] if amount_col < len(row) else None
                if amt_val is not None:
                    try:
                        total_amount += float(amt_val)
                    except (ValueError, TypeError):
                        pass
        
        return order_count, total_amount

    def download_template_from_knowledge_base(self):
        self.log(f"\n[Step 4] 从知识库下载模板文件...")
        
        try:
            list_url = f"{WEKNORA_CONFIG['api_url']}/knowledge-bases/{WEKNORA_CONFIG['knowledge_base_id']}/knowledge"
            response = self._weknora_request('GET', list_url, timeout=10)
            if response is None:
                return None
            
            try:
                result = response.json()
            except Exception as json_err:
                self.log(f"[WARN] 知识库返回解析失败: {json_err}")
                return None
            
            if not result.get('success') or not result.get('data'):
                return None
            
            knowledge_list = result['data']
            template_knowledge = None
            for item in knowledge_list:
                title = item.get('title', '')
                if self.template_file_name in title:
                    template_knowledge = item
                    break
            
            if template_knowledge:
                knowledge_id = template_knowledge['id']
                self.log(f"[OK] 找到模板文件: {template_knowledge['title']}")
                
                download_url = f"{WEKNORA_CONFIG['api_url']}/knowledge/{knowledge_id}/download"
                response = self._weknora_request('GET', download_url, timeout=30)
                if response is None:
                    return None
                
                if response.status_code == 200:
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                    temp_file.write(response.content)
                    temp_file.close()
                    return temp_file.name
                else:
                    return None
            else:
                return None
                
        except Exception as e:
            self.log(f"[WARN] 知识库连接失败: {e}")
        return None

    def load_or_create_workbook(self, template_path):
        if template_path:
            try:
                wb = load_workbook(template_path)
                self.log(f"[OK] 加载模板成功，现有sheet数量: {len(wb.sheetnames)}")
                return wb
            except Exception as e:
                self.log(f"[ERROR] 加载模板失败: {e}")
        
        self.log(f"[INFO] 创建新的工作簿")
        wb = Workbook()
        wb.remove(wb.active)
        return wb

    def update_summary_sheet(self, wb, daily_data):
        self.log(f"\n[Step 5] 更新店铺单量总表...")
        
        sheet_name = '店铺单量总表'
        
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        sub_header_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
        sub_header_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin', color='B4B4B4'),
            right=Side(style='thin', color='B4B4B4'),
            top=Side(style='thin', color='B4B4B4'),
            bottom=Side(style='thin', color='B4B4B4')
        )
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name, index=0)
        
        existing_shops = []
        if ws.max_row > 0 and ws.max_column > 0:
            for col in range(4, ws.max_column + 1, 3):
                shop_name = ws.cell(row=1, column=col).value
                if shop_name:
                    existing_shops.append(shop_name)
        
        if self.is_custom_shops:
            shops_to_keep = self.managed_shops
            shops_to_delete = [shop for shop in existing_shops if shop not in shops_to_keep]
            
            if shops_to_delete:
                self.log(f"  [INFO] 清理汇总表中不需要的店铺列: {shops_to_delete}")
                
                cols_to_delete = []
                for col in range(4, ws.max_column + 1, 3):
                    shop_name = ws.cell(row=1, column=col).value
                    if shop_name in shops_to_delete:
                        cols_to_delete.extend([col, col + 1, col + 2])
                
                cols_to_delete.sort(reverse=True)
                for col in cols_to_delete:
                    ws.delete_cols(col)
                
                existing_shops = [shop for shop in existing_shops if shop not in shops_to_delete]
        
        is_first_time = ws.cell(row=1, column=1).value is None
        if is_first_time:
            cell = ws.cell(row=1, column=1, value='日期')
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            ws.cell(row=2, column=1, value='')
            
            ws.cell(row=1, column=2, value='汇总')
            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
            cell = ws.cell(row=1, column=2)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            cell = ws.cell(row=2, column=2, value='订单量')
            cell.fill = sub_header_fill
            cell.font = sub_header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            cell = ws.cell(row=2, column=3, value='净销售额')
            cell.fill = sub_header_fill
            cell.font = sub_header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 12
        
        new_shops = [shop for shop in self.managed_shops if shop not in existing_shops]
        if new_shops:
            if is_first_time:
                next_col = 4
            else:
                last_shop_col = 0
                for col in range(4, ws.max_column + 1, 3):
                    if ws.cell(row=1, column=col).value:
                        last_shop_col = col
                next_col = last_shop_col + 3 if last_shop_col > 0 else 4
            
            for shop in new_shops:
                ws.cell(row=1, column=next_col, value=shop)
                ws.merge_cells(start_row=1, start_column=next_col, 
                              end_row=1, end_column=next_col + 2)
                cell = ws.cell(row=1, column=next_col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                sub_headers = ['订单量', '净销售额', '净付费率']
                for i, header in enumerate(sub_headers):
                    cell = ws.cell(row=2, column=next_col + i, value=header)
                    cell.fill = sub_header_fill
                    cell.font = sub_header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                ws.column_dimensions[get_column_letter(next_col)].width = 10
                ws.column_dimensions[get_column_letter(next_col + 1)].width = 12
                ws.column_dimensions[get_column_letter(next_col + 2)].width = 10
                
                next_col += 3
        
        date_row = None
        target_date = self.date_yesterday
        for row in range(3, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val:
                # Strict date string match
                row_str = str(cell_val).strip()
                if ' ' in row_str:
                    row_str = row_str.split()[0]
                
                # Handle YYYY-M-D vs YYYY-MM-DD
                try:
                    from datetime import datetime
                    parsed = datetime.strptime(row_str, "%Y-%m-%d")
                    row_normalized = parsed.strftime("%Y-%m-%d")
                except ValueError:
                    try:
                        parsed = datetime.strptime(row_str, "%Y/%m/%d")
                        row_normalized = parsed.strftime("%Y-%m-%d")
                    except ValueError:
                        row_normalized = row_str
                
                if row_normalized == target_date:
                    date_row = row
                    break
        
        if date_row is None:
            date_row = ws.max_row + 1
        
        ws.cell(row=date_row, column=1, value=self.date_yesterday)
        
        # Clear ALL data for this row (except date) to prevent stale data
        for c in range(2, ws.max_column + 1):
            ws.cell(row=date_row, column=c).value = None

        # Recalculate totals strictly from daily_data
        total_orders = 0
        total_net_sales = 0.0
        for data in daily_data.values():
            total_orders += (data.get('order_count', 0) or 0)
            total_net_sales += (data.get('net_sales', 0) or 0)

        total_net_sales = sum(data.get('net_sales', 0) or 0 for data in daily_data.values())
        
        ws.cell(row=date_row, column=2, value=total_orders)
        ws.cell(row=date_row, column=3, value=round(total_net_sales, 2))
        
        from datetime import datetime
        rows_data = []
        for row_idx in range(3, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value
            if date_val:
                rows_data.append((row_idx, date_val))
        
        if len(rows_data) > 1:
            try:
                rows_data.sort(key=lambda x: datetime.strptime(str(x[1]), "%Y-%m-%d"))
                sorted_indices = [x[0] for x in rows_data]
                
                all_data = []
                for row_idx in sorted_indices:
                    row_values = []
                    for col_idx in range(1, ws.max_column + 1):
                        row_values.append(ws.cell(row=row_idx, column=col_idx).value)
                    all_data.append(row_values)
                
                for i, row_values in enumerate(all_data):
                    for j, value in enumerate(row_values):
                        ws.cell(row=3 + i, column=1 + j, value=value)
            except (ValueError, TypeError):
                pass

        # Sorting can move the target date; write shop columns to its new row.
        date_row = next(
            (
                row_idx
                for row_idx in range(3, ws.max_row + 1)
                if str(ws.cell(row=row_idx, column=1).value).strip().split()[0] == self.date_yesterday
            ),
            date_row,
        )
        
        shop_col_map = {}
        # First, identify all shop columns
        for col in range(4, ws.max_column + 1, 3):
            shop_name = ws.cell(row=1, column=col).value
            if shop_name:
                # Strip whitespace to ensure matching with daily_data keys
                shop_col_map[shop_name.strip()] = col
        
        # Force clear all potential shop columns for this date row to prevent stale data
        # We assume max possible columns is 50 to be safe, or use ws.max_column
        limit = max(ws.max_column, 20)
        for col in range(4, limit + 1):
            ws.cell(row=date_row, column=col).value = None

        for shop_name, col in shop_col_map.items():
            # Robust matching with stripping
            match_key = None
            for k in daily_data:
                if k.strip() == shop_name.strip():
                    match_key = k
                    break
            
            if match_key:
                data = daily_data[match_key]
                ws.cell(row=date_row, column=col, value=data.get('order_count', 0))
                ws.cell(row=date_row, column=col + 1, value=round(data.get('net_sales', 0), 2))
                ws.cell(row=date_row, column=col + 2, value=round(data.get('net_pay_ratio', 0), 2))
            else:
                ws.cell(row=date_row, column=col, value=0)
                ws.cell(row=date_row, column=col + 1, value=0)
                ws.cell(row=date_row, column=col + 2, value=0)
        
        self.log(f"[OK] 店铺单量总表更新完成")

    def update_shop_sheet(self, wb, shop_name, shop_data):
        self.log(f"[Step 6] 更新店铺sheet: {shop_name}")
        
        if shop_name in wb.sheetnames:
            ws = wb[shop_name]
        else:
            ws = wb.create_sheet(title=shop_name)
        
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        sub_header_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
        sub_header_font = Font(bold=True, size=10)
        border = Border(
            left=Side(style='thin', color='B4B4B4'),
            right=Side(style='thin', color='B4B4B4'),
            top=Side(style='thin', color='B4B4B4'),
            bottom=Side(style='thin', color='B4B4B4')
        )
        
        new_headers = ['日期', '访客数', '订单量', '成交金额', '净销售额', 
                      '推广费', '付费率', '净付费率', '支出率', 
                      '运费险', '评价排名', '是否剔除补单数据', 
                      '补单金额', '补单数量']
        
        col_widths = [12, 10, 10, 12, 12, 10, 10, 10, 10, 10, 10, 15, 12, 10]
        
        if ws.max_row <= 1:
            ws.cell(row=1, column=1, value=shop_name)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(new_headers))
            title_cell = ws.cell(row=1, column=1)
            title_cell.font = Font(bold=True, size=14, color='1F4E79')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for col_idx, header in enumerate(new_headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            for col_idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        else:
            existing_headers = []
            for col in range(1, ws.max_column + 1):
                header_val = ws.cell(row=2, column=col).value
                if header_val:
                    existing_headers.append(str(header_val))
            
            if len(existing_headers) < len(new_headers):
                self.log(f"[INFO] 检测到旧格式表头，正在扩展为新格式")
                for col_idx, header in enumerate(new_headers, start=1):
                    if col_idx > len(existing_headers) or existing_headers[col_idx - 1] != header:
                        cell = ws.cell(row=2, column=col_idx, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                        if col_idx <= len(col_widths):
                            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
                
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(new_headers))
        
        date_row = None
        is_supplement_removed = False
        for row in range(3, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == self.date_yesterday:
                date_row = row
                is_supplement_removed_val = ws.cell(row=row, column=12).value
                if is_supplement_removed_val is not None:
                    is_supplement_removed = bool(is_supplement_removed_val)
                break
        
        if date_row is None:
            date_row = ws.max_row + 1
        
        shop_data['is_supplement_removed'] = is_supplement_removed
        shop_data['supplement_total_amount'] = shop_data.get('supplement_amount', 0.0)
        shop_data['supplement_order_count'] = shop_data.get('supplement_count', 0)
        
        if not is_supplement_removed and shop_data.get('supplement_count', 0) > 0:
            shop_data['is_supplement_removed'] = True
            self.log(f"[INFO] {shop_name} 在 {self.date_yesterday} 的补单数据已标记为已剔除")
        
        ws.cell(row=date_row, column=1, value=shop_data.get('date', self.date_yesterday))
        ws.cell(row=date_row, column=2, value=shop_data.get('visitor_count', 0))
        ws.cell(row=date_row, column=3, value=shop_data.get('order_count', 0))
        ws.cell(row=date_row, column=4, value=round(shop_data.get('sales_amount', 0), 2))
        ws.cell(row=date_row, column=5, value=round(shop_data.get('net_sales', 0), 2))
        ws.cell(row=date_row, column=6, value=round(shop_data.get('promotion_fee', 0), 2))
        ws.cell(row=date_row, column=7, value=round(shop_data.get('pay_ratio', 0), 2))
        ws.cell(row=date_row, column=8, value=round(shop_data.get('net_pay_ratio', 0), 2))
        ws.cell(row=date_row, column=9, value=round(shop_data.get('daily_expense_rate', 0), 2))
        ws.cell(row=date_row, column=10, value=shop_data.get('service_fee', 0))
        ws.cell(row=date_row, column=11, value=shop_data.get('score_rank', ''))
        ws.cell(row=date_row, column=12, value=shop_data.get('is_supplement_removed', False))
        ws.cell(row=date_row, column=13, value=round(shop_data.get('supplement_total_amount', 0), 2))
        ws.cell(row=date_row, column=14, value=shop_data.get('supplement_order_count', 0))
        
        from datetime import datetime
        rows_data = []
        for row_idx in range(3, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value
            if date_val:
                rows_data.append((row_idx, date_val))
        
        if len(rows_data) > 1:
            try:
                rows_data.sort(key=lambda x: datetime.strptime(str(x[1]), "%Y-%m-%d"))
                sorted_indices = [x[0] for x in rows_data]
                
                all_data = []
                for row_idx in sorted_indices:
                    row_values = []
                    for col_idx in range(1, ws.max_column + 1):
                        row_values.append(ws.cell(row=row_idx, column=col_idx).value)
                    all_data.append(row_values)
                
                for i, row_values in enumerate(all_data):
                    for j, value in enumerate(row_values):
                        ws.cell(row=3 + i, column=1 + j, value=value)
            except (ValueError, TypeError):
                pass

    def delete_old_file_from_knowledge_base(self):
        self.log(f"\n[Step 7] 删除知识库中的旧文件...")
        
        try:
            list_url = f"{WEKNORA_CONFIG['api_url']}/knowledge-bases/{WEKNORA_CONFIG['knowledge_base_id']}/knowledge"
            response = self._weknora_request('GET', list_url, timeout=10)
            if response is None:
                return
            
            try:
                result = response.json()
            except Exception as json_err:
                self.log(f"[WARN] 知识库返回解析失败: {json_err}")
                return
            
            if not result.get('success') or not result.get('data'):
                return
            
            knowledge_list = result['data']
            deleted_count = 0
            for item in knowledge_list:
                title = item.get('title', '')
                if self.output_file_name in title:
                    knowledge_id = item['id']
                    delete_url = f"{WEKNORA_CONFIG['api_url']}/knowledge/{knowledge_id}"
                    
                    del_response = self._weknora_request('DELETE', delete_url, timeout=10)
                    if del_response and del_response.status_code == 200:
                        self.log(f"[OK] 删除旧文件成功: {title}")
                        deleted_count += 1
            
        except Exception as e:
            self.log(f"[WARN] 删除旧文件失败: {e}")

    def upload_to_knowledge_base(self, file_path):
        self.log(f"\n[Step 8] 上传文件到知识库...")
        
        try:
            with open(file_path, 'rb') as f:
                files = {'file': (self.output_file_name, f, 
                                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {'enable_multimodel': 'false'}
                
                upload_url = f"{WEKNORA_CONFIG['api_url']}/knowledge-bases/{WEKNORA_CONFIG['knowledge_base_id']}/knowledge/file"
                response = self._weknora_request('POST', upload_url, files=files, data=data, timeout=30)
            
            if response is None:
                return None
            
            try:
                result = response.json()
            except:
                return 'unknown'
            
            if response.status_code == 200 and result.get('success'):
                file_id = result.get('data', {}).get('id', 'unknown')
                self.log(f"[OK] 文件上传成功，file_id: {file_id}")
                return file_id
            else:
                return None
        except Exception as e:
            self.log(f"[ERROR] 文件上传异常: {e}")
            return None

    def send_to_feishu(self, file_path):
        self.log(f"\n[Step 9] 发送文件到飞书...")
        
        url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
        payload = {"app_id": FEISHU_CONFIG['app_id'], "app_secret": FEISHU_CONFIG['app_secret']}
        response = requests.post(url, json=payload)
        result = response.json()
        
        if result.get('code') != 0:
            self.log(f"[ERROR] 获取飞书token失败: {result}")
            return None
        
        access_token = result['tenant_access_token']
        headers = {'Authorization': f'Bearer {access_token}'}
        
        with open(file_path, 'rb') as f:
            file_content = f.read()
        
        upload_url = "https://open.feishu.cn/open-apis/im/v1/files"
        files = {
            'file_type': (None, 'xlsx'),
            'file_name': (None, self.output_file_name),
            'file': (self.output_file_name, file_content, 
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        response = requests.post(upload_url, headers=headers, files=files)
        upload_result = response.json()
        
        if upload_result.get('code') != 0:
            self.log(f"[ERROR] 飞书文件上传失败: {upload_result}")
            return None
        
        file_key = upload_result['data']['file_key']
        
        send_url = f"https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type={FEISHU_CONFIG['receive_id_type']}"
        payload = {
            "receive_id": FEISHU_CONFIG['receive_id'],
            "msg_type": "file",
            "content": json.dumps({"file_key": file_key}, ensure_ascii=False)
        }
        
        response = requests.post(send_url, headers=headers, json=payload)
        send_result = response.json()
        
        if send_result.get('code') != 0:
            self.log(f"[ERROR] 飞书消息发送失败: {send_result}")
            return None
        
        message_id = send_result['data']['message_id']
        self.log(f"[OK] 飞书发送成功，message_id: {message_id}")
        return message_id

    def save_workbook(self, wb):
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        output_file = os.path.join(OUTPUT_DIR, self.output_file_name)
        wb.save(output_file)
        self.log(f"[OK] 文件保存成功: {output_file}")
        return output_file

    def write_log(self, daily_data, output_file, weknora_file_id, feishu_message_id):
        os.makedirs(LOG_DIR, exist_ok=True)
        log_file = os.path.join(LOG_DIR, f'{datetime.now().strftime("%Y-%m-%d")}.md')
        
        status = "[OK] 成功" if (weknora_file_id and feishu_message_id) else "[PARTIAL] 部分成功"
        
        log_content = f"""## {datetime.now().strftime('%H:%M')} - 每日店铺数据报表 API

### {status}

- **数据日期**: {self.date_yesterday}
- **店铺数量**: {len(daily_data)} 个
- **生成文件**: {output_file}
- **知识库上传**: {'成功' if weknora_file_id else '失败'} (ID: {weknora_file_id if weknora_file_id else 'N/A'})
- **飞书发送**: {'成功' if feishu_message_id else '失败'} (ID: {feishu_message_id if feishu_message_id else 'N/A'})
- **角色**: {self.role}
- **用户**: {self.username if self.username else 'N/A'}

#### 店铺数据摘要
"""
        for shop_name, data in daily_data.items():
            log_content += f"- {shop_name}: 订单 {data.get('order_count', 0)}, 净销售 {data.get('net_sales', 0)}, 每日净支出率 {data.get('daily_expense_rate', 0)}%\n"
        
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(log_content + '\n\n')
        
        self.log(f"[OK] 日志已记录: {log_file}")

    def run(self, upload_to_knowledge_base: bool = True, send_to_feishu: bool = True):
        self.log("=" * 80)
        self.log("每日店铺数据报表 V2.0 - API 模式")
        self.log("=" * 80)
        self.log(f"\n角色: {self.role}")
        if self.username:
            self.log(f"用户名: {self.username}")
        self.log(f"数据日期: {self.date_yesterday}")
        
        local_template_path = os.path.join(OUTPUT_DIR, self.output_file_name)
        template_path = None
        
        if os.path.exists(local_template_path):
            self.log(f"[INFO] 找到本地已存在的报表文件: {local_template_path}")
            template_path = local_template_path
        else:
            self.log(f"[INFO] 本地不存在报表文件，尝试从知识库下载")
            template_path = self.download_template_from_knowledge_base()
        
        wb = self.load_or_create_workbook(template_path)
        
        supplement_removed_status = self.get_supplement_removed_status(wb)
        self.log(f"[INFO] 读取到 {len(supplement_removed_status)} 条补单处理状态记录")
        
        daily_data = self.query_daily_data(supplement_removed_status)
        
        if self.is_custom_shops:
            self.log(f"\n[Step 4.5] 清理不需要的店铺Sheet...")
            sheets_to_keep = ['店铺单量总表'] + self.managed_shops
            sheets_to_delete = []
            for sheet_name in wb.sheetnames:
                if sheet_name not in sheets_to_keep:
                    sheets_to_delete.append(sheet_name)
            
            for sheet_name in sheets_to_delete:
                del wb[sheet_name]
                self.log(f"  [INFO] 删除Sheet: {sheet_name}")
            
            self.log(f"  [OK] 保留Sheet: {wb.sheetnames}")
        
        self.update_summary_sheet(wb, daily_data)
        
        for shop_name in self.managed_shops:
            shop_data = daily_data.get(shop_name, {
                'date': self.date_yesterday,
                'visitor_count': 0,
                'order_count': 0,
                'sales_amount': 0,
                'net_sales': 0,
                'promotion_fee': 0,
                'pay_ratio': 0,
                'net_pay_ratio': 0,
                'daily_expense_rate': 0,
                'service_fee': 0,
                'score_rank': '',
                'supplement_count': 0,
                'supplement_amount': 0,
            })
            self.update_shop_sheet(wb, shop_name, shop_data)
        
        output_file = self.save_workbook(wb)
        
        weknora_file_id = None
        if upload_to_knowledge_base:
            self.delete_old_file_from_knowledge_base()
            weknora_file_id = self.upload_to_knowledge_base(output_file)
        
        feishu_message_id = None
        if send_to_feishu:
            feishu_message_id = self.send_to_feishu(output_file)
        
        if template_path and template_path != local_template_path and os.path.exists(template_path):
            os.unlink(template_path)
        
        self.write_log(daily_data, output_file, weknora_file_id, feishu_message_id)
        
        self.log("\n" + "=" * 80)
        self.log("任务执行完成！")
        self.log("=" * 80)
        
        return {
            'daily_data': daily_data,
            'output_file': output_file,
            'weknora_file_id': weknora_file_id,
            'feishu_message_id': feishu_message_id,
            'logs': self.log_messages
        }


@app.get("/")
async def root():
    return {
        "message": "每日店铺数据报表 API",
        "version": "2.0.0",
        "docs": "/docs"
    }


@app.get("/health")
async def health_check():
    return {"status": "healthy"}


@app.post("/api/report/generate", response_model=ReportResponse)
async def generate_report(request: ReportRequest):
    try:
        reporter = DailyStoreReportAPI(
            role=request.role,
            username=request.username,
            test_date=request.test_date,
            supplement_days_range=request.supplement_days_range
        )
        
        result = reporter.run(
            upload_to_knowledge_base=request.upload_to_knowledge_base,
            send_to_feishu=request.send_to_feishu
        )
        
        return ReportResponse(
            success=True,
            message="报表生成成功",
            data_date=reporter.date_yesterday,
            shop_count=len(result['daily_data']),
            output_file=result['output_file'],
            weknora_file_id=result['weknora_file_id'],
            feishu_message_id=result['feishu_message_id'],
            shop_data=result['daily_data']
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/report/download")
async def download_report(username: str = Query(..., description="用户名")):
    output_file_name = f"{username}的店铺数据日报.xlsx"
    output_file = os.path.join(OUTPUT_DIR, output_file_name)
    if os.path.exists(output_file):
        return FileResponse(
            path=output_file,
            filename=output_file_name,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        raise HTTPException(status_code=404, detail=f"报表文件不存在（用户：{username}），请先生成报表")


@app.post("/api/supplement/upload", response_model=SupplementUploadResponse)
async def upload_supplement_file(
    file: UploadFile = File(...),
    data_date: Optional[str] = Form(default=None, description="补单数据日期，格式：YYYY-MM-DD（不填则从文件名自动提取）"),
    shop_name: Optional[str] = Form(default=None, description="店铺名称（不填则从文件名自动提取）"),
    auto_generate_report: bool = Form(default=True, description="上传后是否自动生成报表"),
    role: str = Form(default="operator", description="角色: boss/supervisor/operator"),
    username: Optional[str] = Form(default="运营主管", description="用户名"),
    upload_to_knowledge_base: bool = Form(default=True, description="报表是否上传到知识库"),
    send_to_feishu: bool = Form(default=True, description="报表是否发送到飞书"),
    supplement_days_range: int = Form(default=3, description="补单文件查找天数范围")
):
    try:
        import re
        
        if not file.filename:
            raise HTTPException(status_code=400, detail="文件名不能为空")
        
        file_ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if file_ext not in ['xls', 'xlsx']:
            raise HTTPException(status_code=400, detail="只支持 .xls 或 .xlsx 格式的文件")
        
        parsed_date = data_date
        parsed_shop = shop_name
        
        if not parsed_date or not parsed_shop:
            filename_match = re.match(
                r'^(\d{4}-\d{1,2}-\d{1,2})(.*?)(?:\s+\d+单)?\.(?:xls|xlsx)$',
                file.filename,
                re.IGNORECASE
            )
            if filename_match:
                if not parsed_date:
                    date_str = filename_match.group(1)
                    try:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        parsed_date = date_obj.strftime('%Y-%m-%d')
                    except ValueError:
                        try:
                            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                            parsed_date = date_obj.strftime('%Y-%m-%d')
                        except ValueError:
                            pass
                if not parsed_shop:
                    parsed_shop = filename_match.group(2).strip()
        
        if not parsed_date:
            raise HTTPException(status_code=400, detail="无法从文件名提取日期，请手动提供 data_date 参数")
        if not parsed_shop:
            raise HTTPException(status_code=400, detail="无法从文件名提取店铺名，请手动提供 shop_name 参数")
        
        data_date = parsed_date
        shop_name = parsed_shop
        
        contents = await file.read()
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_ext}')
        temp_file.write(contents)
        temp_file.close()
        
        try:
            count, amount = 0, 0.0
            is_xls = file_ext == 'xls'
            
            if is_xls and xlrd is None:
                raise RuntimeError("xlrd 未安装，无法读取 .xls 文件")
            
            rows = []
            if is_xls:
                book = xlrd.open_workbook(temp_file.name)
                sheet = book.sheet_by_index(0)
                for r in range(sheet.nrows):
                    row = []
                    for c in range(sheet.ncols):
                        row.append(sheet.cell_value(r, c))
                    rows.append(row)
            else:
                wb = load_workbook(temp_file.name, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    rows.append(list(row))
                wb.close()
            
            if rows:
                header_row_idx = None
                order_col = None
                amount_col = None
                amount_keywords_exact = ['订单金额', '用户实付金额']
                amount_keywords_fuzzy = ['金额']
                
                for idx in range(min(3, len(rows))):
                    row = rows[idx]
                    str_row = [str(v) if v is not None else '' for v in row]
                    if '订单号' in str_row:
                        header_row_idx = idx
                        order_col = str_row.index('订单号')
                        for ci, val in enumerate(str_row):
                            if ci == order_col:
                                continue
                            if any(kw in val for kw in amount_keywords_exact):
                                amount_col = ci
                                break
                        if amount_col is None:
                            for ci, val in enumerate(str_row):
                                if ci == order_col:
                                    continue
                                if any(kw in val for kw in amount_keywords_fuzzy):
                                    amount_col = ci
                                    break
                        break
                
                if header_row_idx is not None and order_col is not None and amount_col is not None:
                    for r in range(header_row_idx + 1, len(rows)):
                        row = rows[r]
                        order_id = row[order_col] if order_col < len(row) else None
                        if order_id is not None and str(order_id).strip():
                            count += 1
                            amt_val = row[amount_col] if amount_col < len(row) else None
                            if amt_val is not None:
                                try:
                                    amount += float(amt_val)
                                except (ValueError, TypeError):
                                    pass
        
        except Exception as e:
            os.unlink(temp_file.name)
            raise HTTPException(status_code=400, detail=f"解析补单文件失败: {str(e)}")
        
        os.unlink(temp_file.name)
        
        new_file_name = f"{data_date}{shop_name} {count}单.{file_ext}"
        
        weknora_file_id = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_ext}') as f:
                f.write(contents)
                temp_upload_path = f.name
            
            try:
                with open(temp_upload_path, 'rb') as f:
                    files = {'file': (new_file_name, f, 
                                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                    data = {'enable_multimodel': 'false'}
                    
                    upload_url = f"{WEKNORA_CONFIG['api_url']}/knowledge-bases/{WEKNORA_CONFIG['knowledge_base_id']}/knowledge/file"
                    
                    def _weknora_upload(method, url, **kwargs):
                        headers = kwargs.pop('headers', {})
                        headers['X-API-Key'] = WEKNORA_CONFIG['api_key']
                        headers['Accept'] = 'application/json'
                        kwargs.setdefault('timeout', 30)
                        
                        _proxy_vars = ['http_proxy', 'https_proxy', 'HTTP_PROXY', 'HTTPS_PROXY', 'ALL_PROXY', 'all_proxy']
                        saved = {}
                        for var in _proxy_vars:
                            if var in os.environ:
                                saved[var] = os.environ.pop(var)
                        
                        try:
                            response = requests.request(method, url, headers=headers, **kwargs)
                            response.encoding = 'utf-8'
                            return response
                        except requests.RequestException as e:
                            return None
                        finally:
                            os.environ.update(saved)
                    
                    response = _weknora_upload('POST', upload_url, files=files, data=data, timeout=30)
                    
                    if response and response.status_code == 200:
                        try:
                            result = response.json()
                            if result.get('success'):
                                weknora_file_id = result.get('data', {}).get('id', 'unknown')
                        except:
                            pass
            finally:
                os.unlink(temp_upload_path)
        except Exception as e:
            pass
        
        report_result = None
        report_generated = False
        
        if auto_generate_report:
            try:
                reporter = DailyStoreReportAPI(
                    role=role,
                    username=username,
                    test_date=data_date,
                    supplement_days_range=supplement_days_range
                )
                
                result = reporter.run(
                    upload_to_knowledge_base=upload_to_knowledge_base,
                    send_to_feishu=send_to_feishu
                )
                
                report_result = ReportResponse(
                    success=True,
                    message="报表生成成功",
                    data_date=reporter.date_yesterday,
                    shop_count=len(result['daily_data']),
                    output_file=result['output_file'],
                    weknora_file_id=result['weknora_file_id'],
                    feishu_message_id=result['feishu_message_id'],
                    shop_data=result['daily_data']
                )
                report_generated = True
            except Exception as e:
                pass
        
        return SupplementUploadResponse(
            success=True,
            message=f"补单文件上传成功！共 {count} 单，金额 {amount:.2f} 元",
            file_name=new_file_name,
            weknora_file_id=weknora_file_id,
            report_generated=report_generated,
            report_result=report_result
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/supplement/batch-upload", response_model=BatchSupplementUploadResponse)
async def batch_upload_supplement_files(
    files: List[UploadFile] = File(...),
    auto_generate_report: bool = Form(default=True, description="上传后是否自动生成报表"),
    role: str = Form(default="operator", description="角色: boss/supervisor/operator"),
    username: Optional[str] = Form(default="运营主管", description="用户名"),
    upload_to_knowledge_base: bool = Form(default=True, description="报表是否上传到知识库"),
    send_to_feishu: bool = Form(default=True, description="报表是否发送到飞书"),
    supplement_days_range: int = Form(default=3, description="补单文件查找天数范围")
):
    try:
        results = []
        success_count = 0
        failed_count = 0
        dates_to_generate = set()
        
        for file in files:
            try:
                import re
                
                if not file.filename:
                    results.append(SupplementUploadResponse(
                        success=False,
                        message="文件名不能为空",
                        file_name="unknown"
                    ))
                    failed_count += 1
                    continue
                
                file_ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
                if file_ext not in ['xls', 'xlsx']:
                    results.append(SupplementUploadResponse(
                        success=False,
                        message=f"只支持 .xls 或 .xlsx 格式的文件，当前: {file_ext}",
                        file_name=file.filename
                    ))
                    failed_count += 1
                    continue
                
                parsed_date = None
                parsed_shop = None
                
                filename_match = re.match(
                    r'^(\d{4}-\d{1,2}-\d{1,2})(.*?)(?:\s+\d+单)?\.(?:xls|xlsx)$',
                    file.filename,
                    re.IGNORECASE
                )
                if filename_match:
                    date_str = filename_match.group(1)
                    try:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        parsed_date = date_obj.strftime('%Y-%m-%d')
                    except ValueError:
                        try:
                            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                            parsed_date = date_obj.strftime('%Y-%m-%d')
                        except ValueError:
                            pass
                    parsed_shop = filename_match.group(2).strip()
                
                if not parsed_date or not parsed_shop:
                    results.append(SupplementUploadResponse(
                        success=False,
                        message="无法从文件名提取日期或店铺名",
                        file_name=file.filename
                    ))
                    failed_count += 1
                    continue
                
                data_date = parsed_date
                shop_name = parsed_shop
                dates_to_generate.add(data_date)
                
                contents = await file.read()
                
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_ext}')
                temp_file.write(contents)
                temp_file.close()
                
                count, amount = 0, 0.0
                try:
                    is_xls = file_ext == 'xls'
                    
                    if is_xls and xlrd is None:
                        raise RuntimeError("xlrd 未安装，无法读取 .xls 文件")
                    
                    rows = []
                    if is_xls:
                        book = xlrd.open_workbook(temp_file.name)
                        sheet = book.sheet_by_index(0)
                        for r in range(sheet.nrows):
                            row = []
                            for c in range(sheet.ncols):
                                row.append(sheet.cell_value(r, c))
                            rows.append(row)
                    else:
                        wb = load_workbook(temp_file.name, data_only=True)
                        ws = wb.active
                        for row in ws.iter_rows(values_only=True):
                            rows.append(list(row))
                        wb.close()
                    
                    if rows:
                        header_row_idx = None
                        order_col = None
                        amount_col = None
                        amount_keywords_exact = ['订单金额', '用户实付金额']
                        amount_keywords_fuzzy = ['金额']
                        
                        for idx in range(min(3, len(rows))):
                            row = rows[idx]
                            str_row = [str(v) if v is not None else '' for v in row]
                            if '订单号' in str_row:
                                header_row_idx = idx
                                order_col = str_row.index('订单号')
                                for ci, val in enumerate(str_row):
                                    if ci == order_col:
                                        continue
                                    if any(kw in val for kw in amount_keywords_exact):
                                        amount_col = ci
                                        break
                                if amount_col is None:
                                    for ci, val in enumerate(str_row):
                                        if ci == order_col:
                                            continue
                                        if any(kw in val for kw in amount_keywords_fuzzy):
                                            amount_col = ci
                                            break
                                break
                        
                        if header_row_idx is not None and order_col is not None and amount_col is not None:
                            for r in range(header_row_idx + 1, len(rows)):
                                row = rows[r]
                                order_id = row[order_col] if order_col < len(row) else None
                                if order_id is not None and str(order_id).strip():
                                    count += 1
                                    amt_val = row[amount_col] if amount_col < len(row) else None
                                    if amt_val is not None:
                                        try:
                                            amount += float(amt_val)
                                        except (ValueError, TypeError):
                                            pass
                except Exception as e:
                    os.unlink(temp_file.name)
                    results.append(SupplementUploadResponse(
                        success=False,
                        message=f"解析补单文件失败: {str(e)}",
                        file_name=file.filename
                    ))
                    failed_count += 1
                    continue
                
                os.unlink(temp_file.name)
                
                new_file_name = f"{data_date}{shop_name} {count}单.{file_ext}"
                
                weknora_file_id = None
                try:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_ext}') as f:
                        f.write(contents)
                        temp_upload_path = f.name
                    
                    try:
                        with open(temp_upload_path, 'rb') as f:
                            files_upload = {'file': (new_file_name, f, 
                                            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                            data_upload = {'enable_multimodel': 'false'}
                            
                            upload_url = f"{WEKNORA_CONFIG['api_url']}/knowledge-bases/{WEKNORA_CONFIG['knowledge_base_id']}/knowledge/file"
                            
                            def _weknora_upload(method, url, **kwargs):
                                headers = kwargs.pop('headers', {})
                                headers['X-API-Key'] = WEKNORA_CONFIG['api_key']
                                headers['Accept'] = 'application/json'
                                kwargs.setdefault('timeout', 30)
                                
                                _proxy_vars = ['http_proxy', 'https_proxy', 'HTTP_PROXY', 'HTTPS_PROXY', 'ALL_PROXY', 'all_proxy']
                                saved = {}
                                for var in _proxy_vars:
                                    if var in os.environ:
                                        saved[var] = os.environ.pop(var)
                                
                                try:
                                    response = requests.request(method, url, headers=headers, **kwargs)
                                    response.encoding = 'utf-8'
                                    return response
                                except requests.RequestException as e:
                                    return None
                                finally:
                                    os.environ.update(saved)
                            
                            response = _weknora_upload('POST', upload_url, files=files_upload, data=data_upload, timeout=30)
                            
                            if response and response.status_code == 200:
                                try:
                                    result_upload = response.json()
                                    if result_upload.get('success'):
                                        weknora_file_id = result_upload.get('data', {}).get('id', 'unknown')
                                except:
                                    pass
                    finally:
                        os.unlink(temp_upload_path)
                except Exception as e:
                    pass
                
                results.append(SupplementUploadResponse(
                    success=True,
                    message=f"补单文件上传成功！共 {count} 单，金额 {amount:.2f} 元",
                    file_name=new_file_name,
                    weknora_file_id=weknora_file_id,
                    report_generated=False
                ))
                success_count += 1
                
            except Exception as e:
                results.append(SupplementUploadResponse(
                    success=False,
                    message=f"处理文件失败: {str(e)}",
                    file_name=file.filename if hasattr(file, 'filename') else "unknown"
                ))
                failed_count += 1
        
        report_result = None
        final_report_generated = False
        
        if auto_generate_report and dates_to_generate:
            try:
                for data_date in sorted(dates_to_generate):
                    reporter = DailyStoreReportAPI(
                        role=role,
                        username=username,
                        test_date=data_date,
                        supplement_days_range=supplement_days_range
                    )
                    
                    result = reporter.run(
                        upload_to_knowledge_base=upload_to_knowledge_base,
                        send_to_feishu=send_to_feishu
                    )
                    
                    report_result = ReportResponse(
                        success=True,
                        message=f"报表生成成功（{data_date}）",
                        data_date=reporter.date_yesterday,
                        shop_count=len(result['daily_data']),
                        output_file=result['output_file'],
                        weknora_file_id=result['weknora_file_id'],
                        feishu_message_id=result['feishu_message_id'],
                        shop_data=result['daily_data']
                    )
                    
                    for res in results:
                        if res.success and hasattr(res, 'file_name') and data_date in res.file_name:
                            res.report_generated = True
                            res.report_result = report_result
                    
                    final_report_generated = True
            except Exception as e:
                pass
        
        return BatchSupplementUploadResponse(
            success=True,
            message=f"批量上传完成：成功 {success_count} 个，失败 {failed_count} 个",
            total_files=len(files),
            success_count=success_count,
            failed_count=failed_count,
            results=results
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/shop-list/upload", response_model=ShopListUploadResponse)
async def upload_shop_list(
    file: UploadFile = File(...),
    data_date: str = Form(..., description="报表日期，格式：YYYY-MM-DD"),
    auto_generate_report: bool = Form(default=True, description="上传后是否自动生成报表"),
    upload_to_knowledge_base: bool = Form(default=True, description="报表是否上传到知识库"),
    send_to_feishu: bool = Form(default=True, description="报表是否发送到飞书"),
    supplement_days_range: int = Form(default=3, description="补单文件查找天数范围")
):
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="文件名不能为空")
        
        file_ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if file_ext not in ['xls', 'xlsx']:
            raise HTTPException(status_code=400, detail="只支持 .xls 或 .xlsx 格式的文件")
        
        contents = await file.read()
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_ext}')
        temp_file.write(contents)
        temp_file.close()
        
        shops = []
        try:
            is_xls = file_ext == 'xls'
            
            if is_xls and xlrd is None:
                raise RuntimeError("xlrd 未安装，无法读取 .xls 文件")
            
            rows = []
            if is_xls:
                book = xlrd.open_workbook(temp_file.name)
                sheet = book.sheet_by_index(0)
                for r in range(sheet.nrows):
                    row = []
                    for c in range(sheet.ncols):
                        row.append(sheet.cell_value(r, c))
                    rows.append(row)
            else:
                wb = load_workbook(temp_file.name, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    rows.append(list(row))
                wb.close()
            
            for row in rows:
                if row:
                    shop_name = str(row[0]).strip()
                    if shop_name and shop_name.lower() not in ['店铺', '店铺名', 'shop', 'shop_name', '店铺名称']:
                        shops.append(shop_name)
            
            shops = list(set(shops))
            shops.sort()
            
        except Exception as e:
            os.unlink(temp_file.name)
            raise HTTPException(status_code=400, detail=f"解析店铺列表文件失败: {str(e)}")
        
        os.unlink(temp_file.name)
        
        if not shops:
            raise HTTPException(status_code=400, detail="未从文件中解析到任何店铺")
        
        report_result = None
        report_generated = False
        
        if auto_generate_report:
            try:
                reporter = DailyStoreReportAPI(
                    role="operator",
                    username=None,
                    test_date=data_date,
                    supplement_days_range=supplement_days_range,
                    custom_shops=shops
                )
                
                result = reporter.run(
                    upload_to_knowledge_base=upload_to_knowledge_base,
                    send_to_feishu=send_to_feishu
                )
                
                report_result = ReportResponse(
                    success=True,
                    message="报表生成成功",
                    data_date=reporter.date_yesterday,
                    shop_count=len(result['daily_data']),
                    output_file=result['output_file'],
                    weknora_file_id=result['weknora_file_id'],
                    feishu_message_id=result['feishu_message_id'],
                    shop_data=result['daily_data']
                )
                report_generated = True
            except Exception as e:
                pass
        
        return ShopListUploadResponse(
            success=True,
            message=f"店铺列表解析成功！共 {len(shops)} 个店铺",
            shop_count=len(shops),
            shops=shops,
            report_generated=report_generated,
            report_result=report_result
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class ShopListWithSupplementUploadResponse(BaseModel):
    success: bool
    message: str
    shop_count: int
    shops: List[str]
    supplement_file_count: int
    report_generated: bool = False
    report_result: Optional[ReportResponse] = None


@app.post("/api/shop-list-with-supplement/upload", response_model=ShopListWithSupplementUploadResponse)
async def upload_shop_list_with_supplement(
    username: Optional[str] = Form(None, description="用户名（可选，可通过店铺列表文件名自动提取）"),
    shop_list_file: UploadFile = File(..., description="店铺列表文件（文件名格式：{username}-负责店铺列表.xlsx）"),
    supplement_files: List[UploadFile] = File(default=[], description="补单数据文件（可多选，可选）"),
    data_date: Optional[str] = Form(None, description="报表日期，格式：YYYY-MM-DD（不填默认昨天）"),
    auto_generate_report: bool = Form(default=True, description="上传后是否自动生成报表"),
    upload_to_knowledge_base: bool = Form(default=False, description="报表是否上传到知识库"),
    send_to_feishu: bool = Form(default=False, description="报表是否发送到飞书")
):
    try:
        if not shop_list_file.filename:
            raise HTTPException(status_code=400, detail="店铺列表文件名不能为空")
        
        # 从文件名提取 username（格式：{username}-负责店铺列表.xlsx）
        extracted_username = None
        if shop_list_file.filename:
            import re
            filename_match = re.match(
                r'^([^-]+)-负责店铺列表\.(?:xls|xlsx)$',
                shop_list_file.filename,
                re.IGNORECASE
            )
            if filename_match:
                extracted_username = filename_match.group(1).strip()
                print(f"[INFO] 从文件名提取到用户名: {extracted_username}")
        
        # 优先使用传入的 username，如果没有则使用从文件名提取的
        final_username = username or extracted_username
        if not final_username:
            raise HTTPException(
                status_code=400, 
                detail="无法获取用户名！请传入 username 参数，或使用文件名格式：{username}-负责店铺列表.xlsx"
            )
        
        shop_list_ext = shop_list_file.filename.rsplit('.', 1)[-1].lower() if '.' in shop_list_file.filename else ''
        if shop_list_ext not in ['xls', 'xlsx']:
            raise HTTPException(status_code=400, detail="店铺列表只支持 .xls 或 .xlsx 格式的文件")
        
        shop_list_contents = await shop_list_file.read()
        
        shop_list_temp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{shop_list_ext}')
        shop_list_temp.write(shop_list_contents)
        shop_list_temp.close()
        
        shops = []
        try:
            is_xls = shop_list_ext == 'xls'
            
            if is_xls and xlrd is None:
                raise RuntimeError("xlrd 未安装，无法读取 .xls 文件")
            
            rows = []
            if is_xls:
                book = xlrd.open_workbook(shop_list_temp.name)
                sheet = book.sheet_by_index(0)
                for r in range(sheet.nrows):
                    row = []
                    for c in range(sheet.ncols):
                        row.append(sheet.cell_value(r, c))
                    rows.append(row)
            else:
                wb = load_workbook(shop_list_temp.name, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    rows.append(list(row))
                wb.close()
            
            for row in rows:
                if row:
                    shop_name = str(row[0]).strip()
                    if shop_name and shop_name.lower() not in ['店铺', '店铺名', 'shop', 'shop_name', '店铺名称']:
                        shops.append(shop_name)
            
            shops = list(set(shops))
            shops.sort()
            
        except Exception as e:
            os.unlink(shop_list_temp.name)
            raise HTTPException(status_code=400, detail=f"解析店铺列表文件失败: {str(e)}")
        
        os.unlink(shop_list_temp.name)
        
        if not shops:
            raise HTTPException(status_code=400, detail="未从文件中解析到任何店铺")
        
        if not data_date:
            yesterday = datetime.now() - timedelta(days=1)
            data_date = yesterday.strftime('%Y-%m-%d')
            print(f"[INFO] 未指定报表日期，使用昨天: {data_date}")
        
        supplement_count = 0
        custom_supplement_data = {}
        
        for supplement_file in supplement_files:
            try:
                if not supplement_file.filename:
                    continue
                
                supp_ext = supplement_file.filename.rsplit('.', 1)[-1].lower() if '.' in supplement_file.filename else ''
                if supp_ext not in ['xls', 'xlsx']:
                    continue
                
                supp_contents = await supplement_file.read()
                
                parsed_date = None
                parsed_shop = None
                
                import re
                filename_match = re.match(
                    r'^(\d{4}-\d{1,2}-\d{1,2})(.*?)(?:\s+\d+单)?\.(?:xls|xlsx)$',
                    supplement_file.filename,
                    re.IGNORECASE
                )
                if filename_match:
                    date_str = filename_match.group(1)
                    try:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        parsed_date = date_obj.strftime('%Y-%m-%d')
                    except ValueError:
                        pass
                    parsed_shop = filename_match.group(2).strip()
                
                if not parsed_date or not parsed_shop:
                    continue
                
                if parsed_date != data_date:
                    print(f"[INFO] 补单文件日期 {parsed_date} 与报表日期 {data_date} 不一致，跳过")
                    continue
                
                count, amount = 0, 0.0
                temp_supp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{supp_ext}')
                temp_supp.write(supp_contents)
                temp_supp.close()
                
                try:
                    is_supp_xls = supp_ext == 'xls'
                    
                    if is_supp_xls and xlrd is None:
                        continue
                    
                    supp_rows = []
                    if is_supp_xls:
                        book = xlrd.open_workbook(temp_supp.name)
                        sheet = book.sheet_by_index(0)
                        for r in range(sheet.nrows):
                            row = []
                            for c in range(sheet.ncols):
                                row.append(sheet.cell_value(r, c))
                            supp_rows.append(row)
                    else:
                        wb = load_workbook(temp_supp.name, data_only=True)
                        ws = wb.active
                        for row in ws.iter_rows(values_only=True):
                            supp_rows.append(list(row))
                        wb.close()
                    
                    if supp_rows:
                        header_row_idx = None
                        order_col = None
                        amount_col = None
                        amount_keywords_exact = ['订单金额', '用户实付金额']
                        amount_keywords_fuzzy = ['金额']
                        
                        for idx in range(min(3, len(supp_rows))):
                            row = supp_rows[idx]
                            str_row = [str(v) if v is not None else '' for v in row]
                            if '订单号' in str_row:
                                header_row_idx = idx
                                order_col = str_row.index('订单号')
                                for ci, val in enumerate(str_row):
                                    if ci == order_col:
                                        continue
                                    if any(kw in val for kw in amount_keywords_exact):
                                        amount_col = ci
                                        break
                                if amount_col is None:
                                    for ci, val in enumerate(str_row):
                                        if ci == order_col:
                                            continue
                                        if any(kw in val for kw in amount_keywords_fuzzy):
                                            amount_col = ci
                                            break
                                break
                        
                        if header_row_idx is not None and order_col is not None and amount_col is not None:
                            for r in range(header_row_idx + 1, len(supp_rows)):
                                row = supp_rows[r]
                                order_id = row[order_col] if order_col < len(row) else None
                                if order_id is not None and str(order_id).strip():
                                    count += 1
                                    amt_val = row[amount_col] if amount_col < len(row) else None
                                    if amt_val is not None:
                                        try:
                                            amount += float(amt_val)
                                        except (ValueError, TypeError):
                                            pass
                    
                    if count > 0:
                        custom_supplement_data[parsed_shop] = {
                            'count': count,
                            'amount': amount
                        }
                        print(f"[OK] 解析补单文件: {supplement_file.filename} -> {parsed_shop}: {count}单, {amount}元")
                    
                    supplement_count += 1
                    
                except Exception as e:
                    print(f"[ERROR] 解析补单文件失败: {e}")
                finally:
                    if os.path.exists(temp_supp.name):
                        os.unlink(temp_supp.name)
                
            except Exception as e:
                continue
        
        report_result = None
        report_generated = False
        
        if auto_generate_report:
            try:
                reporter = DailyStoreReportAPI(
                    role="operator",
                    username=final_username,
                    test_date=data_date,
                    supplement_days_range=0,
                    custom_shops=shops,
                    custom_supplement_data=custom_supplement_data
                )
                
                result = reporter.run(
                    upload_to_knowledge_base=upload_to_knowledge_base,
                    send_to_feishu=send_to_feishu
                )
                
                report_result = ReportResponse(
                    success=True,
                    message="报表生成成功",
                    data_date=reporter.date_yesterday,
                    shop_count=len(result['daily_data']),
                    output_file=result['output_file'],
                    weknora_file_id=result['weknora_file_id'],
                    feishu_message_id=result['feishu_message_id'],
                    shop_data=result['daily_data']
                )
                report_generated = True
            except Exception as e:
                print(f"[ERROR] 生成报表失败: {e}")
                import traceback
                traceback.print_exc()
        
        return ShopListWithSupplementUploadResponse(
            success=True,
            message=f"处理完成！店铺 {len(shops)} 个，补单文件 {supplement_count} 个",
            shop_count=len(shops),
            shops=shops,
            supplement_file_count=supplement_count,
            report_generated=report_generated,
            report_result=report_result
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
